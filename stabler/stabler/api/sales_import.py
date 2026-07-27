import hashlib
import io
import json
from datetime import date, datetime

import frappe
import openpyxl
from frappe import _
from frappe.utils import flt, nowtime, today

from stabler.api._common import _require_company
from stabler.api.approvals import _assert_company_scope

ITEM_CODE_REMAP = {
	"5": "005",
	"15": "15-17",
	"17": "15-17",
	"61": "61-62",
}


def parse_number(value) -> float:
	if value is None:
		return 0.0
	if isinstance(value, (int, float)):
		return float(value)
	cleaned = str(value).replace("\xa0", "").replace(" ", "").strip()
	if not cleaned or cleaned == "-":
		return 0.0
	try:
		return float(cleaned)
	except ValueError:
		return 0.0


def parse_date(value) -> str:
	if isinstance(value, datetime):
		return value.date().isoformat()
	if isinstance(value, date):
		return value.isoformat()
	text = str(value).strip()
	date_part = text.split(" ")[0].split("T")[0]
	for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
		try:
			return datetime.strptime(date_part, fmt).date().isoformat()
		except ValueError:
			continue
	raise ValueError(f"Unrecognized date value: {value!r}")


def read_excel(file_content) -> list[dict]:
	wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
	ws = wb.active

	header_row = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
	col = {h: i for i, h in enumerate(header_row) if h}

	required = ("date", "customer", "item", "box", "box kg", "rate")
	missing = [r for r in required if r not in col]
	if missing:
		raise ValueError(f"Missing required columns: {', '.join(missing)}")

	rows = []
	for row in ws.iter_rows(min_row=2, values_only=True):
		if not row[col["date"]]:
			continue
		item_raw = str(row[col["item"]]).strip() if row[col.get("item")] else ""
		if not item_raw or item_raw.lower() in ("none", ""):
			continue

		row_dict = {
			"date": parse_date(row[col["date"]]),
			"item_code": ITEM_CODE_REMAP.get(item_raw, item_raw),
			"boxes": parse_number(row[col["box"]]),
			"box_kg": parse_number(row[col["box kg"]]),
			"total_kg": parse_number(row[col.get("total kg", col.get("box"))]),
			"rate": parse_number(row[col["rate"]]),
			"amount": parse_number(row[col.get("amount", col.get("rate"))]),
			"customer": str(row[col["customer"]]).strip() if row[col.get("customer")] else "",
			"parent": str(row[col["parent"]]).strip()
			if "parent" in col and row[col["parent"]] is not None
			else "",
		}
		rows.append(row_dict)

	wb.close()
	return rows


def _effective_customer(row: dict) -> str:
	whom = (row.get("whom") or "").strip()
	if whom:
		return whom
	return (row.get("customer") or "").strip()


def auto_correct_total_kg(rows: list[dict]) -> None:
	for r in rows:
		amount = r.get("amount") or 0.0
		rate = r.get("rate") or 0.0
		total_kg = r.get("total_kg") or 0.0
		if amount > 0 and rate > 0 and total_kg > 0 and abs(amount - total_kg * rate) > 0.5:
			r["original_total_kg"] = total_kg
			r["total_kg"] = round(amount / rate, 3)
			r["total_kg_corrected"] = True


def auto_correct_boxes(rows: list[dict]) -> None:
	for r in rows:
		box_kg = r.get("box_kg") or 0.0
		total_kg = r.get("total_kg") or 0.0
		boxes = r.get("boxes") or 0.0
		if box_kg > 0 and total_kg > 0 and abs(boxes * box_kg - total_kg) > 0.5:
			r["original_boxes"] = boxes
			r["boxes"] = round(total_kg / box_kg, 2)
			r["boxes_corrected"] = True


def resolve_customers(rows: list[dict]) -> tuple[dict[str, str], list[str]]:
	names = sorted({_effective_customer(r) for r in rows if _effective_customer(r)})
	resolved = {}
	missing = []
	for name in names:
		cust = frappe.db.get_value("Customer", {"customer_name": name}, "name")
		if not cust:
			cust = frappe.db.get_value("Customer", {"name": name}, "name")
		if cust:
			resolved[name] = cust
		else:
			missing.append(name)
	return resolved, missing


def resolve_items(rows: list[dict]) -> tuple[set[str], list[str]]:
	codes = sorted({r["item_code"] for r in rows if r["item_code"]})
	known = set()
	missing = []
	for code in codes:
		if frappe.db.exists("Item", code):
			known.add(code)
		else:
			missing.append(code)
	return known, missing


def _import_ref(payload: dict) -> str:
	items_canon = sorted((i["item_code"], i["qty"], i["rate"]) for i in payload.get("items", []))
	canonical = json.dumps(
		{
			"customer": payload["customer"],
			"posting_date": payload["posting_date"],
			"is_return": payload.get("is_return", 0),
			"items": items_canon,
		},
		sort_keys=True,
	)
	return "IMPORT-" + hashlib.sha1(canonical.encode()).hexdigest()[:12]


def build_payloads(rows: list[dict], customer_map: dict[str, str], company: str) -> list[dict]:
	from collections import defaultdict

	groups = defaultdict(list)
	for r in rows:
		key = (r["date"], _effective_customer(r))
		groups[key].append(r)

	payloads = []
	for key, g_rows in sorted(groups.items(), key=lambda kv: kv[0]):
		posting_date, customer_display = key[0], key[1]
		erpnext_name = customer_map.get(customer_display)
		if not erpnext_name:
			continue

		base = {
			"customer": erpnext_name,
			"company": company,
			"set_posting_time": 1,
			"posting_date": posting_date,
			"currency": "UZS",
			"update_stock": 0,
		}

		def _is_positive(r):
			if r["boxes"] != 0:
				return r["boxes"] > 0
			if r["total_kg"] != 0:
				return r["total_kg"] > 0
			return r.get("amount", 0) > 0

		def _is_negative(r):
			if r["boxes"] != 0:
				return r["boxes"] < 0
			if r["total_kg"] != 0:
				return r["total_kg"] < 0
			return r.get("amount", 0) < 0

		pos_rows = [r for r in g_rows if _is_positive(r)]
		neg_rows = [r for r in g_rows if _is_negative(r)]

		if pos_rows:
			pos_payload = {
				**base,
				"items": [
					{
						"item_code": r["item_code"],
						"qty": abs(r["total_kg"]),
						"rate": r["rate"],
						"uom": "Kg",
						"custom_boxes": abs(r["boxes"]),
						"custom_box_kg": r["box_kg"],
						"allow_zero_valuation_rate": 1,
						**({"is_free_item": 1, "discount_percentage": 100} if r["rate"] == 0 else {}),
					}
					for r in pos_rows
				],
			}
			pos_payload["po_no"] = _import_ref(pos_payload)
			payloads.append(pos_payload)

		if neg_rows:
			neg_payload = {
				**base,
				"is_return": 1,
				"items": [
					{
						"item_code": r["item_code"],
						"qty": -abs(r["total_kg"]),
						"rate": r["rate"],
						"uom": "Kg",
						"custom_boxes": -abs(r["boxes"]),
						"custom_box_kg": r["box_kg"],
						"allow_zero_valuation_rate": 1,
						**({"is_free_item": 1, "discount_percentage": 100} if r["rate"] == 0 else {}),
					}
					for r in neg_rows
				],
			}
			neg_payload["po_no"] = _import_ref(neg_payload)
			payloads.append(neg_payload)

	return payloads


def _get_file_content(file_url=None):
	if "file" in frappe.request.files:
		return frappe.request.files["file"].stream.read()
	if file_url:
		file_name = frappe.db.get_value("File", {"file_url": file_url}, "name")
		if not file_name or not frappe.has_permission("File", "read", file_name):
			frappe.throw(_("Not permitted to access this file."), frappe.PermissionError)
		from frappe.utils.file_manager import get_file

		_file_name, content = get_file(file_url)
		return content
	frappe.throw(_("No file uploaded or file URL provided."))


@frappe.whitelist()
def preview_sales_import(file_url=None, corrections=None):
	content = _get_file_content(file_url)
	rows = read_excel(content)

	if corrections:
		if isinstance(corrections, str):
			corrections = json.loads(corrections)
		for row_str, corr in corrections.items():
			row_idx = int(row_str) - 2
			if 0 <= row_idx < len(rows):
				row = rows[row_idx]
				if "boxes" in corr:
					row["boxes"] = float(corr["boxes"])
				if "box_kg" in corr:
					row["box_kg"] = float(corr["box_kg"])
				if "rate" in corr:
					row["rate"] = float(corr["rate"])

	auto_correct_total_kg(rows)
	auto_correct_boxes(rows)

	_customer_map, missing_customers = resolve_customers(rows)
	_item_codes, missing_items = resolve_items(rows)

	issues = []
	for c in missing_customers:
		issues.append(
			{
				"severity": "blocker",
				"code": "customer_missing",
				"message": f"Customer not found in accounting system: '{c}'",
				"customer": c,
			}
		)
	for code in missing_items:
		issues.append(
			{
				"severity": "blocker",
				"code": "item_missing",
				"message": f"Item code not found in accounting system: '{code}'",
				"item": code,
			}
		)

	for idx, r in enumerate(rows, start=2):
		if not r.get("customer") and not r.get("parent"):
			issues.append(
				{
					"severity": "blocker",
					"code": "customer_blank",
					"message": "Customer/buyer name is empty",
					"row": idx,
				}
			)
		if r["rate"] < 0:
			issues.append(
				{
					"severity": "blocker",
					"code": "rate_invalid",
					"message": "Rate cannot be negative",
					"row": idx,
				}
			)
		if r["total_kg"] == 0 and r.get("amount", 0) == 0:
			issues.append(
				{
					"severity": "blocker",
					"code": "qty_zero",
					"message": "Quantity and amount are zero",
					"row": idx,
				}
			)

	blocker_count = sum(1 for i in issues if i["severity"] == "blocker")
	return {
		"row_count": len(rows),
		"blocker_count": blocker_count,
		"issues": issues,
		"can_import": blocker_count == 0,
	}


@frappe.whitelist()
def execute_sales_import(file_url=None, corrections=None, selected_indices=None, company=None):
	if not frappe.has_permission("Sales Invoice", "create"):
		frappe.throw(_("You are not permitted to import sales invoices."), frappe.PermissionError)

	if not company:
		company = (
			frappe.defaults.get_user_default("Company") or frappe.get_all("Company", pluck="name", limit=1)[0]
		)
	_require_company(company)
	_assert_company_scope(company)

	content = _get_file_content(file_url)
	rows = read_excel(content)

	if corrections:
		if isinstance(corrections, str):
			corrections = json.loads(corrections)
		for row_str, corr in corrections.items():
			row_idx = int(row_str) - 2
			if 0 <= row_idx < len(rows):
				row = rows[row_idx]
				if "boxes" in corr:
					row["boxes"] = float(corr["boxes"])
				if "box_kg" in corr:
					row["box_kg"] = float(corr["box_kg"])
				if "rate" in corr:
					row["rate"] = float(corr["rate"])

	auto_correct_total_kg(rows)
	auto_correct_boxes(rows)

	customer_map, missing_customers = resolve_customers(rows)
	_item_codes, missing_items = resolve_items(rows)

	if missing_customers or missing_items:
		frappe.throw("Cannot import: please fix missing items/customers first.")

	payloads = build_payloads(rows, customer_map, company)

	if selected_indices:
		if isinstance(selected_indices, str):
			selected_indices = json.loads(selected_indices)
		payloads = [p for idx, p in enumerate(payloads) if idx in selected_indices]

	created = []
	errors = []
	skipped = []

	for payload in payloads:
		ref = payload.get("po_no")
		if ref and frappe.db.exists("Sales Invoice", {"po_no": ref, "docstatus": 1}):
			skipped.append({"customer": payload["customer"], "reason": "Already imported"})
			continue

		try:
			doc = frappe.new_doc("Sales Invoice")
			doc.update(payload)
			doc.insert()
			doc.submit()
			created.append(doc.name)
		except Exception as exc:
			frappe.db.rollback()
			errors.append({"customer": payload["customer"], "error": str(exc)})

	return {
		"status": "success",
		"created_count": len(created),
		"error_count": len(errors),
		"skipped_count": len(skipped),
		"created": created,
		"errors": errors,
		"skipped": skipped,
	}
