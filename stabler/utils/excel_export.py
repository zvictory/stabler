"""Professional, reusable Excel (.xlsx) writer for Stabler reports.

No frappe import — pure openpyxl + stdlib, so it is unit-testable without a bench
and reusable by any export endpoint. The Frappe layer (stabler/api/export.py)
fetches permission-checked data and hands it here; this module only formats.

Input model mirrors the `_shape()` output used across stabler/api/reports.py:
  columns: [{"key","label","type","align"}]   type ∈ text|int|number|money|percent|date
  rows:    [ {col_key: value, ...}, ... ]
  totals:  { col_key: value, ... }             (partial — only summed columns)
  header_meta: ordered [(label, value), ...]   rendered as a metadata block on top

Layout (top → bottom):
  1  Title  (merged, bold, large)
  2..k Metadata block (Company / Date range / Currency / Generated / Filters / Note)
  blank
  H  Column header row  (bold, filled, frozen, autofilter)
  …  Data rows          (per-type number formats, negative money in red)
  T  Totals row         (bold, top border) for columns present in `totals`
"""

from __future__ import annotations

import datetime
import io
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── House style ──────────────────────────────────────────────────────────────
_TITLE_FONT = Font(bold=True, size=14, color="1F2D3D")
_META_LABEL_FONT = Font(bold=True, size=9, color="667382")
_META_VALUE_FONT = Font(size=9, color="1F2D3D")
_HEADER_FONT = Font(bold=True, size=10, color="1F2D3D")
_HEADER_FILL = PatternFill("solid", fgColor="EEF1F5")
_TOTALS_FONT = Font(bold=True, size=10, color="1F2D3D")
_TOTALS_FILL = PatternFill("solid", fgColor="F6F8FA")
_THIN = Side(style="thin", color="D0D5DB")
_MEDIUM = Side(style="medium", color="AAB2BD")
_HEADER_BORDER = Border(bottom=_MEDIUM)
_TOTALS_BORDER = Border(top=_MEDIUM)
_CELL_BORDER = Border(bottom=_THIN)

# Number formats. Negatives in red for money/number so losses/credits stand out.
_FMT_MONEY = "#,##0.00;[Red]-#,##0.00"
_FMT_INT = "#,##0;[Red]-#,##0"
_FMT_NUMBER = "#,##0.###;[Red]-#,##0.###"
_FMT_PERCENT = '0.0"%"'  # our percent values are already ×100 (e.g. 40.0 → 40.0%)
_FMT_DATE = "dd.mm.yyyy"

_RIGHT = Alignment(horizontal="right")
_LEFT = Alignment(horizontal="left", wrap_text=False)
_CENTER = Alignment(horizontal="center")

_NUMERIC_TYPES = {"int", "number", "money", "percent"}


def safe_sheet_name(name: str) -> str:
	"""Excel sheet names: ≤31 chars, none of []:*?/\\, not blank."""
	s = re.sub(r"[\[\]:*?/\\]", " ", str(name or "Report")).strip()
	return s[:31] or "Report"


def report_filename(prefix: str, date_from=None, date_to=None, suffix=None) -> str:
	"""Professional, descriptive filename (no random `export.xlsx`)."""
	parts = [re.sub(r"[^\w-]+", "_", str(prefix or "Report")).strip("_")]
	if date_from and date_to:
		parts.append(f"{date_from}_to_{date_to}")
	elif date_from:
		parts.append(str(date_from))
	if suffix:
		parts.append(re.sub(r"[^\w-]+", "_", str(suffix)).strip("_"))
	return "_".join(p for p in parts if p) + ".xlsx"


def _fmt_for(col_type: str):
	return {
		"money": _FMT_MONEY,
		"int": _FMT_INT,
		"number": _FMT_NUMBER,
		"percent": _FMT_PERCENT,
		"date": _FMT_DATE,
	}.get(col_type)


def _align_for(col):
	if col.get("align") == "end" or col.get("type") in _NUMERIC_TYPES:
		return _RIGHT
	if col.get("type") == "date":
		return _CENTER
	return _LEFT


def _coerce(value, col_type: str):
	"""Turn a JSON-ish value into something Excel stores natively."""
	if value is None or value == "":
		return None
	if col_type in _NUMERIC_TYPES:
		try:
			return float(value)
		except (TypeError, ValueError):
			return value
	if col_type == "date":
		if isinstance(value, (datetime.date, datetime.datetime)):
			return value
		m = re.match(r"(\d{4})-(\d{2})-(\d{2})", str(value))
		if m:
			return datetime.date(int(m[1]), int(m[2]), int(m[3]))
		return str(value)
	return str(value)


def build_report_workbook(
	*,
	title: str,
	columns: list,
	rows: list,
	totals: dict | None = None,
	header_meta: list | None = None,
	sheet_name: str | None = None,
) -> Workbook:
	"""Build a styled workbook. Returns an openpyxl Workbook (caller serializes)."""
	wb = Workbook()
	ws = wb.active
	ws.title = safe_sheet_name(sheet_name or title)
	ncols = max(1, len(columns))
	last_col_letter = get_column_letter(ncols)

	r = 1
	# ── Title ──
	ws.cell(row=r, column=1, value=str(title)).font = _TITLE_FONT
	if ncols > 1:
		ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
	r += 1

	# ── Metadata block ──
	for label, value in header_meta or []:
		if value in (None, ""):
			continue
		lc = ws.cell(row=r, column=1, value=f"{label}:")
		lc.font = _META_LABEL_FONT
		vc = ws.cell(row=r, column=2, value=str(value))
		vc.font = _META_VALUE_FONT
		if ncols > 2:
			ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=ncols)
		r += 1

	r += 1  # blank spacer
	header_row = r

	# ── Column headers ──
	for c, col in enumerate(columns, start=1):
		cell = ws.cell(row=header_row, column=c, value=str(col.get("label") or col.get("key") or ""))
		cell.font = _HEADER_FONT
		cell.fill = _HEADER_FILL
		cell.border = _HEADER_BORDER
		cell.alignment = _align_for(col)
	r += 1

	# ── Data rows ──
	widths = [len(str(col.get("label") or "")) + 2 for col in columns]
	for row in rows:
		for c, col in enumerate(columns, start=1):
			val = _coerce(row.get(col["key"]), col.get("type"))
			cell = ws.cell(row=r, column=c, value=val)
			fmt = _fmt_for(col.get("type"))
			if fmt:
				cell.number_format = fmt
			cell.alignment = _align_for(col)
			cell.border = _CELL_BORDER
			widths[c - 1] = max(widths[c - 1], len(str(val)) + 2 if val is not None else 0)
		r += 1

	# ── Totals row ──
	totals = totals or {}
	if totals and rows:
		first = True
		for c, col in enumerate(columns, start=1):
			cell = ws.cell(row=r, column=c)
			cell.font = _TOTALS_FONT
			cell.fill = _TOTALS_FILL
			cell.border = _TOTALS_BORDER
			if first and col["key"] not in totals:
				cell.value = "Total"
				cell.alignment = _LEFT
				first = False
				continue
			if col["key"] in totals:
				cell.value = _coerce(totals[col["key"]], col.get("type"))
				fmt = _fmt_for(col.get("type"))
				if fmt:
					cell.number_format = fmt
				cell.alignment = _align_for(col)
				first = False
		r += 1

	# ── Freeze panes (keep title+meta+header on scroll) + autofilter ──
	ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
	last_data_row = r - 1
	ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{max(header_row, last_data_row)}"

	# ── Column widths (capped) ──
	for i, w in enumerate(widths, start=1):
		ws.column_dimensions[get_column_letter(i)].width = min(48, max(10, w))

	return wb


def normalize_frappe_columns(cols) -> list:
	"""Map a Frappe query-report `columns` list into our {key,label,type}.

	Handles modern dict columns ({label, fieldname, fieldtype}) and the legacy
	"Label:Type:Width" string form."""
	out = []
	for c in cols or []:
		if isinstance(c, dict):
			key = c.get("fieldname") or c.get("key") or c.get("label")
			label = c.get("label") or key
			ftype = str(c.get("fieldtype") or "").lower()
		else:
			parts = str(c).split(":")
			label = parts[0]
			key = label
			ftype = (parts[1] if len(parts) > 1 else "").lower()
		if ftype in ("currency", "float"):
			t = "money"
		elif ftype == "int":
			t = "int"
		elif ftype == "percent":
			t = "percent"
		elif ftype in ("date", "datetime"):
			t = "date"
		else:
			t = "text"
		out.append({"key": key, "label": label, "type": t})
	return out


def _row_get(row, key, idx):
	"""Result rows may be dicts (keyed by fieldname) or positional lists."""
	if isinstance(row, dict):
		return row.get(key)
	try:
		return row[idx]
	except (IndexError, TypeError):
		return None


def build_financial_statement_workbook(
	*,
	title: str,
	columns: list,
	rows: list,
	header_meta: list | None = None,
	sheet_name: str | None = None,
	label_key: str | None = None,
	indent_key: str = "indent",
	bold_predicate=None,
) -> Workbook:
	"""Template for hierarchical statements (P&L / Balance Sheet / Trial Balance).

	The first (label) column is indented by each row's `indent` level; section and
	total rows are bold. Numeric columns get the money format with red negatives.
	"""
	wb = Workbook()
	ws = wb.active
	ws.title = safe_sheet_name(sheet_name or title)
	ncols = max(1, len(columns))
	label_key = label_key or (columns[0]["key"] if columns else "label")

	r = 1
	ws.cell(row=r, column=1, value=str(title)).font = _TITLE_FONT
	if ncols > 1:
		ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
	r += 1
	for label, value in header_meta or []:
		if value in (None, ""):
			continue
		ws.cell(row=r, column=1, value=f"{label}:").font = _META_LABEL_FONT
		ws.cell(row=r, column=2, value=str(value)).font = _META_VALUE_FONT
		if ncols > 2:
			ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=ncols)
		r += 1
	r += 1
	header_row = r
	for c, col in enumerate(columns, start=1):
		cell = ws.cell(row=header_row, column=c, value=str(col.get("label") or col.get("key") or ""))
		cell.font = _HEADER_FONT
		cell.fill = _HEADER_FILL
		cell.border = _HEADER_BORDER
		cell.alignment = _align_for(col)
	r += 1

	widths = [len(str(col.get("label") or "")) + 2 for col in columns]
	for row in rows:
		indent = 0
		if isinstance(row, dict):
			try:
				indent = int(row.get(indent_key) or 0)
			except (TypeError, ValueError):
				indent = 0
		is_bold = bool(bold_predicate(row)) if bold_predicate else False
		for c, col in enumerate(columns, start=1):
			raw = _row_get(row, col["key"], c - 1)
			val = _coerce(raw, col.get("type"))
			cell = ws.cell(row=r, column=c, value=val)
			if col["key"] == label_key:
				cell.alignment = Alignment(horizontal="left", indent=min(8, indent))
			else:
				fmt = _fmt_for(col.get("type"))
				if fmt:
					cell.number_format = fmt
				cell.alignment = _align_for(col)
			if is_bold:
				cell.font = _TOTALS_FONT
			cell.border = _CELL_BORDER
			label_len = (len(str(val)) + 2 + indent * 2) if val is not None else 0
			widths[c - 1] = max(widths[c - 1], label_len)
		r += 1

	ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
	for i, w in enumerate(widths, start=1):
		ws.column_dimensions[get_column_letter(i)].width = min(60, max(12, w))
	return wb


def _write_title_meta(ws, title, header_meta, ncols) -> int:
	"""Shared title + metadata block. Returns the header row index."""
	r = 1
	ws.cell(row=r, column=1, value=str(title)).font = _TITLE_FONT
	if ncols > 1:
		ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
	r += 1
	for label, value in header_meta or []:
		if value in (None, ""):
			continue
		ws.cell(row=r, column=1, value=f"{label}:").font = _META_LABEL_FONT
		ws.cell(row=r, column=2, value=str(value)).font = _META_VALUE_FONT
		if ncols > 2:
			ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=ncols)
		r += 1
	return r + 1  # blank spacer → header row


def build_ledger_workbook(
	*,
	title: str,
	entries: list,
	opening: float,
	header_meta: list | None = None,
	sheet_name: str | None = None,
	date_key: str = "posting_date",
	voucher_key: str = "voucher_no",
	debit_key: str = "debit_in_account_currency",
	credit_key: str = "credit_in_account_currency",
	sign: str = "dr",
) -> Workbook:
	"""Party / account ledger: opening balance → movements (debit/credit + running
	balance) → totals → closing balance. `sign`: "dr" → balance = debit−credit
	(receivable); "cr" → credit−debit (payable)."""
	wb = Workbook()
	ws = wb.active
	ws.title = safe_sheet_name(sheet_name or title)
	cols = [
		{"key": "_date", "label": "Date", "type": "date"},
		{"key": "_voucher", "label": "Voucher", "type": "text"},
		{"key": "_debit", "label": "Debit", "type": "money", "align": "end"},
		{"key": "_credit", "label": "Credit", "type": "money", "align": "end"},
		{"key": "_balance", "label": "Balance", "type": "money", "align": "end"},
	]
	ncols = len(cols)
	header_row = _write_title_meta(ws, title, header_meta, ncols)
	for c, col in enumerate(cols, start=1):
		cell = ws.cell(row=header_row, column=c, value=col["label"])
		cell.font = _HEADER_FONT
		cell.fill = _HEADER_FILL
		cell.border = _HEADER_BORDER
		cell.alignment = _align_for(col)
	r = header_row + 1

	def _row(date_v, voucher_v, debit_v, credit_v, balance_v, *, bold=False):
		nonlocal r
		vals = [
			_coerce(date_v, "date"),
			str(voucher_v or ""),
			_coerce(debit_v, "money") if debit_v is not None else None,
			_coerce(credit_v, "money") if credit_v is not None else None,
			_coerce(balance_v, "money") if balance_v is not None else None,
		]
		for c, (col, v) in enumerate(zip(cols, vals, strict=True), start=1):
			cell = ws.cell(row=r, column=c, value=v)
			fmt = _fmt_for(col["type"])
			if fmt:
				cell.number_format = fmt
			cell.alignment = _align_for(col)
			cell.border = _CELL_BORDER
			if bold:
				cell.font = _TOTALS_FONT
				cell.fill = _TOTALS_FILL
		r += 1

	_row(None, "Opening balance", None, None, opening, bold=True)
	bal = float(opening or 0)
	tot_dr = tot_cr = 0.0
	for e in entries:
		dr = float(e.get(debit_key) or 0)
		cr = float(e.get(credit_key) or 0)
		tot_dr += dr
		tot_cr += cr
		bal += (dr - cr) if sign == "dr" else (cr - dr)
		_row(e.get(date_key), e.get(voucher_key), dr or None, cr or None, bal)
	_row(None, "Total", tot_dr, tot_cr, None, bold=True)

	ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
	for i, w in zip(range(1, ncols + 1), (14, 38, 16, 16, 18), strict=True):
		ws.column_dimensions[get_column_letter(i)].width = w
	return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
	buf = io.BytesIO()
	wb.save(buf)
	return buf.getvalue()
