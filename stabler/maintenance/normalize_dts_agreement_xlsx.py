#!/usr/bin/env python3
"""Normalize DTS agreement receivables Excel rows for preview_agreement_opening.

This command only reads the workbook and emits JSON; it never calls ERPNext or
creates financial documents.
"""

from __future__ import annotations

import argparse
import json
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


HEADERS = {
    "organization": {"организация", "organization", "customer"},
    "agreement": {"договор", "agreement", "contract"},
    "amount": {"всего", "total", "amount", "balance"},
}


def _text(value) -> str:
    return str(value or "").strip()


def _header_map(row) -> dict[str, int]:
    normalized = {_text(value).lower(): index for index, value in enumerate(row)}
    result = {}
    for target, aliases in HEADERS.items():
        for alias in aliases:
            if alias in normalized:
                result[target] = normalized[alias]
                break
    return result


def normalize(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    header = None
    header_map = {}
    rows = list(sheet.iter_rows(values_only=True))
    for index, row in enumerate(rows):
        candidate = _header_map(row)
        if set(candidate) == set(HEADERS):
            header = index
            header_map = candidate
            break
    if header is None:
        raise ValueError("Could not find Organization/Agreement/Total columns")

    result = []
    for row_number, row in enumerate(rows[header + 1 :], start=header + 2):
        organization = _text(row[header_map["organization"]])
        agreement = _text(row[header_map["agreement"]])
        raw_amount = row[header_map["amount"]]
        if (not organization and raw_amount in (None, "")) or (not agreement and raw_amount in (None, "", 0, 0.0)):
            continue
        if not organization:
            raise ValueError(f"Row {row_number} requires organization")
        try:
            amount = Decimal(str(raw_amount or 0).replace(" ", "").replace(",", "."))
        except Exception as exc:
            raise ValueError(f"Row {row_number} has invalid amount: {raw_amount!r}") from exc
        result.append({
            "organization": organization,
            "agreement": agreement,
            "amount": float(amount),
            "currency": "UZS",
            "as_of_date": "2026-07-20",
            "source_row": row_number,
        })
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    payload = normalize(args.workbook)
    encoded = json.dumps(payload, ensure_ascii=False, indent=2)
    if args.output:
        args.output.write_text(encoded + "\n", encoding="utf-8")
    else:
        print(encoded)


if __name__ == "__main__":
    main()
