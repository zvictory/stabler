"""Workbook-validation tests for the professional Excel writer (no frappe/DB).

Generates a real .xlsx in memory and inspects it with openpyxl — covers the QA
checklist: title, header, frozen panes, autofilter, number formats, totals row,
Cyrillic preservation, sheet-name + filename rules.
"""

import datetime
import io
import os
import sys
import unittest

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")))

from openpyxl import load_workbook

from stabler.utils.excel_export import (
	build_financial_statement_workbook,
	build_ledger_workbook,
	build_report_workbook,
	normalize_frappe_columns,
	report_filename,
	safe_sheet_name,
	workbook_to_bytes,
)

COLUMNS = [
	{"key": "customer_name", "label": "Клиент", "type": "text"},
	{"key": "invoice_count", "label": "Invoices", "type": "int", "align": "end"},
	{"key": "total", "label": "Sales", "type": "money", "align": "end"},
	{"key": "margin_pct", "label": "Margin %", "type": "percent", "align": "end"},
	{"key": "posting_date", "label": "Date", "type": "date"},
]
ROWS = [
	{"customer_name": "АО Растворобетон", "invoice_count": 3, "total": 1250000.5, "margin_pct": 40.0, "posting_date": "2026-06-12"},
	{"customer_name": "URALSK", "invoice_count": 1, "total": -5000.0, "margin_pct": -12.5, "posting_date": "2026-06-18"},
]
TOTALS = {"invoice_count": 4, "total": 1245000.5}
META = [("Company", "Anjan"), ("Date range", "2026-06-01 → 2026-06-30"), ("Currency", "UZS")]


def _wb():
	wb = build_report_workbook(
		title="Sales by Customer",
		columns=COLUMNS,
		rows=ROWS,
		totals=TOTALS,
		header_meta=META,
		sheet_name="Sales by Customer",
	)
	return load_workbook(io.BytesIO(workbook_to_bytes(wb)))


class TestWorkbook(unittest.TestCase):
	def setUp(self):
		self.wb = _wb()
		self.ws = self.wb.active

	def test_opens_and_sheet_name(self):
		self.assertEqual(self.ws.title, "Sales by Customer")

	def test_title_cell(self):
		self.assertEqual(self.ws["A1"].value, "Sales by Customer")
		self.assertTrue(self.ws["A1"].font.bold)

	def test_metadata_block(self):
		# A2 label, B2 value
		self.assertEqual(self.ws["A2"].value, "Company:")
		self.assertEqual(self.ws["B2"].value, "Anjan")

	def test_header_row_and_freeze_autofilter(self):
		# Title(1) + 3 meta + blank(1) → header on row 6
		header_row = 6
		labels = [self.ws.cell(row=header_row, column=c).value for c in range(1, 6)]
		self.assertEqual(labels, ["Клиент", "Invoices", "Sales", "Margin %", "Date"])
		self.assertTrue(self.ws.cell(row=header_row, column=1).font.bold)
		# Freeze panes start just below the header.
		self.assertEqual(self.ws.freeze_panes, "A7")
		self.assertIsNotNone(self.ws.auto_filter.ref)
		self.assertTrue(self.ws.auto_filter.ref.startswith("A6"))

	def test_number_formats(self):
		# Data starts row 7. Column 3 = money, 4 = percent, 5 = date.
		money = self.ws.cell(row=7, column=3)
		pct = self.ws.cell(row=7, column=4)
		date = self.ws.cell(row=7, column=5)
		self.assertIn("#,##0.00", money.number_format)
		self.assertIn("[Red]", money.number_format)  # negatives styled
		self.assertIn("%", pct.number_format)
		self.assertEqual(date.number_format, "dd.mm.yyyy")
		self.assertIsInstance(date.value, datetime.date)

	def test_cyrillic_preserved(self):
		self.assertEqual(self.ws.cell(row=7, column=1).value, "АО Растворобетон")

	def test_totals_row(self):
		# Rows 7,8 data → totals on row 9.
		self.assertEqual(self.ws.cell(row=9, column=1).value, "Total")
		self.assertEqual(self.ws.cell(row=9, column=2).value, 4)
		self.assertEqual(self.ws.cell(row=9, column=3).value, 1245000.5)
		self.assertTrue(self.ws.cell(row=9, column=2).font.bold)


class TestHelpers(unittest.TestCase):
	def test_safe_sheet_name(self):
		self.assertEqual(safe_sheet_name("Sales/Customer:2026*?"), "Sales Customer 2026")
		self.assertEqual(safe_sheet_name(""), "Report")
		self.assertLessEqual(len(safe_sheet_name("x" * 50)), 31)

	def test_filename(self):
		self.assertEqual(
			report_filename("Sales by Customer", "2026-06-01", "2026-06-30", "Anjan"),
			"Sales_by_Customer_2026-06-01_to_2026-06-30_Anjan.xlsx",
		)
		self.assertEqual(report_filename("Inventory Aging", "2026-06-23"), "Inventory_Aging_2026-06-23.xlsx")


class TestFrappeColumns(unittest.TestCase):
	def test_dict_columns(self):
		cols = normalize_frappe_columns([
			{"label": "Account", "fieldname": "account", "fieldtype": "Link"},
			{"label": "Balance", "fieldname": "balance", "fieldtype": "Currency"},
			{"label": "Qty", "fieldname": "qty", "fieldtype": "Int"},
		])
		self.assertEqual(cols[0], {"key": "account", "label": "Account", "type": "text"})
		self.assertEqual(cols[1]["type"], "money")
		self.assertEqual(cols[2]["type"], "int")

	def test_legacy_string_columns(self):
		cols = normalize_frappe_columns(["Account:Link:200", "Balance:Currency:120"])
		self.assertEqual(cols[1]["type"], "money")
		self.assertEqual(cols[0]["label"], "Account")


class TestFinancialStatement(unittest.TestCase):
	def setUp(self):
		cols = [
			{"key": "account", "label": "Account", "type": "text"},
			{"key": "amount", "label": "2026-06", "type": "money", "align": "end"},
		]
		rows = [
			{"account": "Income", "amount": 1000000, "indent": 0},
			{"account": "Sales", "amount": 1000000, "indent": 1},
			{"account": "Total Income", "amount": 1000000, "indent": 0},
			{"account": "Expense", "amount": -400000, "indent": 0},
		]
		wb = build_financial_statement_workbook(
			title="Profit and Loss",
			columns=cols,
			rows=rows,
			header_meta=[("Company", "Anjan")],
			sheet_name="P&L",
			bold_predicate=lambda r: int(r.get("indent") or 0) == 0 or str(r.get("account", "")).lower().startswith("total"),
		)
		self.ws = load_workbook(io.BytesIO(workbook_to_bytes(wb))).active

	def test_indent_and_bold(self):
		# Title(1)+1 meta+blank → header row 4, data starts row 5.
		# Row 5 = Income (indent 0, bold); row 6 = Sales (indent 1, not bold).
		income = self.ws.cell(row=5, column=1)
		sales = self.ws.cell(row=6, column=1)
		self.assertTrue(income.font.bold)
		self.assertEqual(sales.alignment.indent, 1)
		self.assertFalse(sales.font.bold)

	def test_money_format_and_negative(self):
		expense_amt = self.ws.cell(row=8, column=2)  # Expense row
		self.assertIn("[Red]", expense_amt.number_format)
		self.assertEqual(expense_amt.value, -400000)


class TestLedger(unittest.TestCase):
	def setUp(self):
		entries = [
			{"posting_date": "2026-06-02", "voucher_no": "ACC-SINV-001", "debit_in_account_currency": 1000000, "credit_in_account_currency": 0},
			{"posting_date": "2026-06-10", "voucher_no": "ACC-PAY-001", "debit_in_account_currency": 0, "credit_in_account_currency": 400000},
		]
		wb = build_ledger_workbook(
			title="Customer Ledger — ACME",
			entries=entries,
			opening=200000,
			header_meta=[("Company", "Anjan"), ("Currency", "UZS")],
			sheet_name="Customer Ledger",
			sign="dr",
		)
		self.ws = load_workbook(io.BytesIO(workbook_to_bytes(wb))).active

	def test_opening_running_closing(self):
		# Title(1)+2 meta+blank → header row 5; opening row 6.
		self.assertEqual(self.ws.cell(row=6, column=2).value, "Opening balance")
		self.assertEqual(self.ws.cell(row=6, column=5).value, 200000)
		# After +1,000,000 debit → balance 1,200,000 (row 7).
		self.assertEqual(self.ws.cell(row=7, column=5).value, 1200000)
		# After −400,000 credit → balance 800,000 (row 8).
		self.assertEqual(self.ws.cell(row=8, column=5).value, 800000)
		# Total row (9): debit 1,000,000 / credit 400,000.
		self.assertEqual(self.ws.cell(row=9, column=2).value, "Total")
		self.assertEqual(self.ws.cell(row=9, column=3).value, 1000000)
		self.assertEqual(self.ws.cell(row=9, column=4).value, 400000)

	def test_payable_sign(self):
		wb = build_ledger_workbook(
			title="Supplier Ledger",
			entries=[{"posting_date": "2026-06-01", "voucher_no": "PINV-1", "debit_in_account_currency": 0, "credit_in_account_currency": 500000}],
			opening=0,
			sign="cr",
		)
		ws = load_workbook(io.BytesIO(workbook_to_bytes(wb))).active
		# credit 500,000 with cr sign → balance +500,000 (we owe).
		# header row: title(1)+blank → 3; opening 4; entry 5.
		self.assertEqual(ws.cell(row=5, column=5).value, 500000)


if __name__ == "__main__":
	unittest.main()
