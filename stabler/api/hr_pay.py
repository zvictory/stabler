"""Stabler payroll computation — runs the ported anjan-hr engine over a
`Stabler Payroll Attendance Summary` and returns the full pay + breakdown.

Read-only previews here; emitting ERPNext Additional Salary / Salary Slip stays
in hr_payroll_calc. Salary is sensitive → gated to payroll-visible roles.
"""

from __future__ import annotations

import frappe
from stabler.api.approvals import _assert_company_scope
from frappe import _
from frappe.utils import flt

from stabler.api._common import _require_company
from stabler.api._payroll_adapter import build_calc_input
from stabler.api._payroll_calc import calculate_payroll

_SUMMARY = "Stabler Payroll Attendance Summary"
_RULESET = "Stabler Attendance Rule Set"
_PAY_ROLES = {"HR Manager", "Payroll Manager", "Accounts Manager", "System Manager", "Stabler Admin"}


def _require_pay_role() -> None:
	if frappe.session.user == "Guest":
		frappe.throw(_("Login required."), frappe.PermissionError)
	if not (set(frappe.get_roles()) & _PAY_ROLES):
		frappe.throw(_("You need a payroll/HR role to view computed pay."), frappe.PermissionError)


def _employee_pay_fields(emp_name: str) -> dict:
	e = (
		frappe.db.get_value(
			"Employee",
			emp_name,
			[
				"employee_name", "date_of_joining", "custom_base_salary", "custom_allowance_config",
				"custom_work_mode", "custom_stake_coefficient", "custom_region",
				"custom_heavy_conditions", "custom_additional_duties", "custom_duty_supplement_pct",
			],
			as_dict=True,
		)
		or {}
	)
	return {
		"employee_name": e.get("employee_name"),
		"base_salary": e.get("custom_base_salary") or 0,
		"stake_coefficient": e.get("custom_stake_coefficient") or 1,
		"work_mode": e.get("custom_work_mode") or "SHIFT_8H",
		"region": e.get("custom_region") or "NO_TRAVEL",
		"heavy_conditions": bool(e.get("custom_heavy_conditions")),
		"additional_duties": bool(e.get("custom_additional_duties")),
		"allowance_config": e.get("custom_allowance_config"),
		"hire_date": str(e.get("date_of_joining")) if e.get("date_of_joining") else None,
		"duty_supplement_pct": e.get("custom_duty_supplement_pct") or 0,
	}


def _region_rates(rs: dict) -> dict:
	"""Assemble the engine's region_rates map from the rule set's flat per-band
	Currency fields. Only non-zero bands are included; NO_TRAVEL is always zero
	(handled in the engine)."""
	out = {}
	for band, field in (
		("CITY", "region_city_rate"),
		("DISTRICT", "region_district_rate"),
		("FAR_DISTRICT", "region_far_district_rate"),
	):
		val = rs.get(field)
		if val:
			out[band] = val
	return out


def _ruleset_dict(company: str) -> dict:
	# No cross-company fallback — only the requesting company's ruleset is used.
	name = frappe.db.get_value(_RULESET, {"company": company, "enabled": 1, "is_default": 1}, "name")
	if not name:
		return {}
	rs = frappe.get_doc(_RULESET, name).as_dict()
	# Compose the engine-facing region_rates map from the flat per-band fields.
	rs["region_rates"] = _region_rates(rs)
	return rs


def _duty_supplements(emp: dict) -> list:
	"""Map the per-employee duty-supplement % onto the engine's list input."""
	try:
		pct = float(emp.get("duty_supplement_pct") or 0)
	except (TypeError, ValueError):
		pct = 0.0
	return [{"pct": pct}] if pct else []


def _summary_fields(s) -> dict:
	return {
		"payroll_period": s.payroll_period,
		"present_days": s.present_days,
		"absent_days": s.absent_days,
		"half_days": s.half_days,
		"overtime_minutes": s.overtime_minutes,
		"night_minutes": s.night_minutes,
		"late_deduction_amount": s.late_deduction_amount,
	}


def _advance_adjustments(s) -> list:
	"""Per-period salary-advance recovery → engine ADVANCE adjustment."""
	amt = flt(getattr(s, "advance_deduction", 0) or 0)
	return [{"type": "ADVANCE", "amount": amt}] if amt else []


def _compute(s, ruleset: dict) -> dict:
	emp = _employee_pay_fields(s.employee)
	inp = build_calc_input(
		emp,
		_summary_fields(s),
		ruleset,
		kpi_performance_pct=getattr(s, "kpi_performance_pct", None),
		duty_supplements=_duty_supplements(emp),
		adjustments=_advance_adjustments(s),
	)
	result = calculate_payroll(inp)
	result["summary"] = s.name
	result["employee"] = s.employee
	result["employee_name"] = emp.get("employee_name") or s.employee
	result["period"] = s.payroll_period
	result["status"] = s.status
	result["kpi_performance_pct"] = getattr(s, "kpi_performance_pct", 0) or 0
	result["advance_deduction"] = flt(getattr(s, "advance_deduction", 0) or 0)
	return result


@frappe.whitelist()
def preview_payroll_pay(summary_name: str) -> dict:
	"""Full computed pay + breakdown for one attendance summary (read-only)."""
	_require_pay_role()
	# Fetch company before loading the full doc to prevent IDOR enumeration.
	company = frappe.db.get_value(_SUMMARY, summary_name, "company")
	if not company:
		frappe.throw(_("Unknown summary: {0}").format(summary_name))
	_require_company(company)
	_assert_company_scope(company)  # tenant isolation: reject a foreign company arg
	s = frappe.get_doc(_SUMMARY, summary_name)
	return _compute(s, _ruleset_dict(company))


@frappe.whitelist()
def set_kpi_performance(summary_name: str, pct) -> dict:
	"""Set an employee's KPI performance % (0-100) for a period, then recompute.

	Writes the score onto the attendance summary and returns the freshly computed
	pay so the UI can update in place. Role-gated + company-scoped + audited via
	the doctype's track_changes.
	"""
	_require_pay_role()
	company = frappe.db.get_value(_SUMMARY, summary_name, "company")
	if not company:
		frappe.throw(_("Unknown summary: {0}").format(summary_name))
	_require_company(company)
	_assert_company_scope(company)  # tenant isolation: reject a foreign company arg
	try:
		value = float(pct)
	except (TypeError, ValueError):
		frappe.throw(_("KPI performance must be a number."))
	if value < 0 or value > 100:
		frappe.throw(_("KPI performance must be between 0 and 100."))
	s = frappe.get_doc(_SUMMARY, summary_name)
	if s.status == "Locked":
		frappe.throw(_("This period is locked; KPI cannot be changed."))
	s.db_set("kpi_performance_pct", value)
	s.reload()
	return _compute(s, _ruleset_dict(company))


@frappe.whitelist()
def set_advance_deduction(summary_name: str, amount) -> dict:
	"""Set the salary-advance amount recovered from net pay this period, recompute.

	Capped at the employee's outstanding advance balance so a period can never
	over-recover. Role-gated, company-scoped, blocked on locked periods.
	"""
	_require_pay_role()
	company = frappe.db.get_value(_SUMMARY, summary_name, "company")
	if not company:
		frappe.throw(_("Unknown summary: {0}").format(summary_name))
	_require_company(company)
	_assert_company_scope(company)  # tenant isolation: reject a foreign company arg
	try:
		value = flt(amount)
	except (TypeError, ValueError):
		frappe.throw(_("Advance deduction must be a number."))
	if value < 0:
		frappe.throw(_("Advance deduction cannot be negative."))
	s = frappe.get_doc(_SUMMARY, summary_name)
	if s.status == "Locked":
		frappe.throw(_("This period is locked; the advance deduction cannot be changed."))
	outstanding = _advance_outstanding(company).get(s.employee, 0.0)
	if value > outstanding + 0.005:
		frappe.throw(
			_("Deduction {0} exceeds the outstanding advance balance {1}.").format(value, round(outstanding, 2))
		)
	s.db_set("advance_deduction", value)
	s.reload()
	r = _compute(s, _ruleset_dict(company))
	r["advance_outstanding"] = outstanding
	return r


def _advance_outstanding(company: str) -> dict:
	"""Outstanding advance balance per employee (delegates to employee_advance)."""
	try:
		from stabler.api.employee_advance import _advance_account, _balances

		return _balances(company, _advance_account(company))
	except Exception:
		# No advance account configured → treat as no outstanding advances.
		return {}


@frappe.whitelist()
def preview_payroll_period(company: str, payroll_period: str) -> dict:
	"""Computed pay for every summary in a period — for a payroll review screen."""
	_require_pay_role()
	_require_company(company)
	_assert_company_scope(company)  # tenant isolation: reject a foreign company arg
	ruleset = _ruleset_dict(company)
	outstanding = _advance_outstanding(company)
	names = frappe.get_all(
		_SUMMARY,
		filters={"company": company, "payroll_period": payroll_period},
		pluck="name",
		limit=0,
	)
	rows = []
	gross_total = 0.0
	net_total = 0.0
	for name in names:
		s = frappe.get_doc(_SUMMARY, name)
		if s.company != company:
			continue  # TOCTOU guard — skip if doc moved to another company between query and load
		r = _compute(s, ruleset)
		r["advance_outstanding"] = outstanding.get(s.employee, 0.0)
		gross_total += float(r["breakdown"].get("gross") or 0)
		net_total += float(r.get("net") or 0)
		rows.append(r)
	return {
		"company": company,
		"period": payroll_period,
		"count": len(rows),
		"gross_total": round(gross_total),
		"net_total": round(net_total),
		"rows": rows,
	}


@frappe.whitelist()
def payroll_xlsx(company: str, payroll_period: str):
	"""Colored Excel of the period's computed pay (role-gated download)."""
	_assert_company_scope(company)  # tenant isolation: reject a foreign company arg
	import io

	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Font, PatternFill
	from openpyxl.utils import get_column_letter

	data = preview_payroll_period(company, payroll_period)
	rows = data.get("rows") or []
	base_ccy = frappe.get_cached_value("Company", company, "default_currency") or ""

	wb = Workbook()
	ws = wb.active
	ws.title = "Payroll"
	header_fill = PatternFill("solid", fgColor="206BC4")
	white = Font(bold=True, color="FFFFFF")
	center = Alignment(horizontal="center")
	cols = [
		(_("Employee"), "employee_name", 28),
		(_("Prorated base"), "prorated_base", 16),
		(_("Allowances"), "allowances", 14),
		(_("Overtime"), "overtime", 12),
		(_("KPI"), "kpi", 12),
		(_("Duty supplement"), "duty_supplement", 14),
		(_("Bonus"), "bonus", 12),
		(_("Fines"), "fines", 12),
		(_("Advance"), "advance", 14),
		(_("Net"), "net", 16),
	]
	ws.append([f"{_('Payroll')} · {payroll_period} · {company}"])
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
	ws.cell(row=1, column=1).font = Font(bold=True, size=12)

	hr = 2
	for ci, (label, key, width) in enumerate(cols, start=1):
		c = ws.cell(row=hr, column=ci, value=label)
		c.fill = header_fill
		c.font = white
		c.alignment = center
		ws.column_dimensions[get_column_letter(ci)].width = width

	r = hr + 1
	totals = {key: 0.0 for (label, key, width) in cols if key != "employee_name"}
	for row in rows:
		ws.cell(row=r, column=1, value=row.get("employee_name") or row.get("employee"))
		for ci, (label, key, width) in enumerate(cols[1:], start=2):
			val = flt(row.get(key))
			totals[key] = totals.get(key, 0.0) + val
			cell = ws.cell(row=r, column=ci, value=round(val))
			cell.number_format = "#,##0"
			if key in ("fines", "advance") and val:
				cell.font = Font(color="D63939")
			if key == "net":
				cell.font = Font(bold=True)
		r += 1

	# Totals row.
	tcell = ws.cell(row=r, column=1, value=_("Total"))
	tcell.font = Font(bold=True)
	for ci, (label, key, width) in enumerate(cols[1:], start=2):
		c = ws.cell(row=r, column=ci, value=round(totals.get(key, 0.0)))
		c.number_format = "#,##0"
		c.font = Font(bold=True)
	ws.cell(row=r + 1, column=1, value=f"{_('Amounts in')} {base_ccy}").font = Font(italic=True, size=9, color="888888")
	ws.freeze_panes = "A3"

	buf = io.BytesIO()
	wb.save(buf)
	frappe.local.response.filename = f"payroll-{payroll_period}.xlsx"
	frappe.local.response.filecontent = buf.getvalue()
	frappe.local.response.type = "binary"
