"""HR module — Employees, Attendance, Leave, Payroll basics."""

from __future__ import annotations

import json

import frappe
from frappe.utils import flt, getdate, today, add_days, date_diff


from stabler.api._common import _require_company, _assert_can_read, _assert_can_write


# ---------------------------------------------------------------------------
# Salary-visibility helper
# ---------------------------------------------------------------------------

#: Roles whose holders may read/write salary-sensitive fields.
_PAYROLL_VISIBLE_ROLES = frozenset(
	("Accounts Manager", "Payroll Manager", "HR Manager", "System Manager")
)

#: Fields masked for non-payroll-visible users.
SALARY_FIELDS = frozenset(("custom_base_salary", "custom_allowance_config"))

#: All native + custom employee fields surfaced by Stabler.
_EMPLOYEE_NATIVE_FIELDS = (
	"employee_name",
	"image",
	"department",
	"designation",
	"date_of_joining",
	"relieving_date",
	"status",
	"cell_number",
	"company",
)

_EMPLOYEE_CUSTOM_FIELDS = (
	"custom_timepay_id",
	"custom_timepay_name",
	"custom_base_salary",
	"custom_shift_class",
	"custom_region",
	"custom_work_mode",
	"custom_stake_coefficient",
	"custom_heavy_conditions",
	"custom_additional_duties",
	"custom_duty_supplement_pct",
	"custom_allowance_config",
)

#: All writable field names (whitelist for update_employee).
_EMPLOYEE_WRITABLE_FIELDS = frozenset(
	_EMPLOYEE_NATIVE_FIELDS + _EMPLOYEE_CUSTOM_FIELDS
)

#: Enum validators for custom fields.
_VALID_SHIFT_CLASS = frozenset(("DAY", "NIGHT", "OFFICE", "LIGHT"))
_VALID_REGION = frozenset(("CITY", "DISTRICT", "FAR_DISTRICT", "NO_TRAVEL"))
_VALID_WORK_MODE = frozenset(("SHIFT_8H", "SHIFT_12H", "HALF_RATE", "FLEXIBLE", "REMOTE"))


def _user_can_see_salary(user: str | None = None) -> bool:
	"""Return True when the session/given user holds a payroll-visible role."""
	user = user or frappe.session.user
	if user in ("Administrator",):
		return True
	from stabler.api.organization import _ADMIN_ROLES
	roles = set(frappe.get_roles(user))
	return bool(roles & (_PAYROLL_VISIBLE_ROLES | set(_ADMIN_ROLES)))


def _parse_items(items):
	if items is None:
		return []
	if isinstance(items, str):
		try:
			return json.loads(items)
		except Exception:
			frappe.throw("Invalid items payload (expected JSON).")
	return list(items)


# ----- Employees -----------------------------------------------------------


@frappe.whitelist()
def list_employees(company: str, search: str = "", status: str = "", limit: int = 100):
	_require_company(company)
	conds = ["company = %(company)s"]
	params: dict = {"company": company, "limit": int(limit)}
	if status:
		conds.append("status = %(status)s")
		params["status"] = status
	if search:
		conds.append(
			"(name LIKE %(s)s OR employee_name LIKE %(s)s OR cell_number LIKE %(s)s OR user_id LIKE %(s)s)"
		)
		params["s"] = f"%{search}%"
	where = " AND ".join(conds)
	return frappe.db.sql(
		f"""
		SELECT name, employee_name, status, designation, department,
		       date_of_joining, cell_number, user_id, image, gender
		FROM `tabEmployee`
		WHERE {where}
		ORDER BY status='Active' DESC, employee_name
		LIMIT %(limit)s
		""",
		params,
		as_dict=True,
	)


@frappe.whitelist()
def employee_detail(name: str):
	_assert_can_read("Employee", name)
	if not name or not frappe.db.exists("Employee", name):
		frappe.throw(f"Unknown employee: {name}")
	doc = frappe.get_doc("Employee", name)
	can_see_salary = _user_can_see_salary()
	payload = {
		# Core identity
		"name": doc.name,
		"employee_name": doc.employee_name,
		"first_name": doc.first_name,
		"last_name": doc.last_name,
		"status": doc.status,
		"company": doc.company,
		"image": getattr(doc, "image", None),
		# Role / org
		"department": doc.department,
		"designation": doc.designation,
		# Dates
		"date_of_birth": doc.date_of_birth,
		"date_of_joining": doc.date_of_joining,
		"relieving_date": getattr(doc, "relieving_date", None),
		# Contact
		"gender": doc.gender,
		"cell_number": doc.cell_number,
		"personal_email": getattr(doc, "personal_email", None),
		"company_email": getattr(doc, "company_email", None),
		"user_id": doc.user_id,
		# Misc
		"holiday_list": doc.holiday_list,
		"employment_type": getattr(doc, "employment_type", None),
		# Custom — identity / integration
		"custom_timepay_id": getattr(doc, "custom_timepay_id", None),
		"custom_timepay_name": getattr(doc, "custom_timepay_name", None),
		# Custom — work configuration
		"custom_shift_class": getattr(doc, "custom_shift_class", None),
		"custom_region": getattr(doc, "custom_region", None),
		"custom_work_mode": getattr(doc, "custom_work_mode", None),
		"custom_stake_coefficient": flt(getattr(doc, "custom_stake_coefficient", 1.0)),
		"custom_heavy_conditions": int(getattr(doc, "custom_heavy_conditions", 0) or 0),
		"custom_additional_duties": int(getattr(doc, "custom_additional_duties", 0) or 0),
		"custom_duty_supplement_pct": flt(getattr(doc, "custom_duty_supplement_pct", 0) or 0),
		# Custom — salary (masked for non-payroll roles)
		"custom_base_salary": flt(getattr(doc, "custom_base_salary", 0)) if can_see_salary else None,
		"custom_allowance_config": getattr(doc, "custom_allowance_config", None) if can_see_salary else None,
	}
	return payload


@frappe.whitelist()
def create_employee(
	company: str,
	first_name: str,
	last_name: str = "",
	gender: str = "",
	date_of_birth: str = "",
	date_of_joining: str = "",
	designation: str = "",
	department: str = "",
	cell_number: str = "",
	user_id: str = "",
	# Custom fields (all optional)
	custom_timepay_id: str = "",
	custom_timepay_name: str = "",
	custom_shift_class: str = "",
	custom_region: str = "",
	custom_work_mode: str = "",
	custom_stake_coefficient: float = 1.0,
	custom_heavy_conditions: int = 0,
	custom_additional_duties: int = 0,
	custom_base_salary: float = 0.0,
	custom_allowance_config: str = "",
):
	_require_company(company)
	if not first_name:
		frappe.throw("First name is required.")
	if not gender:
		frappe.throw("Gender is required.")
	if not date_of_birth:
		frappe.throw("Date of birth is required.")
	if not date_of_joining:
		date_of_joining = today()

	# Validate custom enums
	if custom_shift_class and custom_shift_class not in _VALID_SHIFT_CLASS:
		frappe.throw(f"Invalid custom_shift_class: {custom_shift_class}. Must be one of {sorted(_VALID_SHIFT_CLASS)}.")
	if custom_region and custom_region not in _VALID_REGION:
		frappe.throw(f"Invalid custom_region: {custom_region}. Must be one of {sorted(_VALID_REGION)}.")
	if custom_work_mode and custom_work_mode not in _VALID_WORK_MODE:
		frappe.throw(f"Invalid custom_work_mode: {custom_work_mode}. Must be one of {sorted(_VALID_WORK_MODE)}.")

	# Validate stake_coefficient
	coeff = flt(custom_stake_coefficient or 1.0)
	if custom_work_mode != "HALF_RATE":
		coeff = 1.0
	elif not (0.1 <= coeff <= 2.0):
		frappe.throw("custom_stake_coefficient must be between 0.1 and 2.0.")

	# Validate allowance_config JSON if provided
	if custom_allowance_config:
		try:
			json.loads(custom_allowance_config)
		except Exception:
			frappe.throw("custom_allowance_config must be valid JSON.")

	# Salary fields — only payroll-visible users may set them
	can_see_salary = _user_can_see_salary()

	doc = frappe.new_doc("Employee")
	doc.company = company
	doc.first_name = first_name
	doc.last_name = last_name or ""
	doc.gender = gender
	doc.date_of_birth = getdate(date_of_birth)
	doc.date_of_joining = getdate(date_of_joining)
	doc.status = "Active"
	if designation:
		doc.designation = designation
	if department:
		doc.department = department
	if cell_number:
		doc.cell_number = cell_number
	if user_id:
		doc.user_id = user_id
	# Custom fields
	if custom_timepay_id:
		doc.custom_timepay_id = custom_timepay_id
	if custom_timepay_name:
		doc.custom_timepay_name = custom_timepay_name
	if custom_shift_class:
		doc.custom_shift_class = custom_shift_class
	if custom_region:
		doc.custom_region = custom_region
	if custom_work_mode:
		doc.custom_work_mode = custom_work_mode
	doc.custom_stake_coefficient = coeff
	doc.custom_heavy_conditions = int(custom_heavy_conditions or 0)
	doc.custom_additional_duties = int(custom_additional_duties or 0)
	if can_see_salary:
		if flt(custom_base_salary):
			doc.custom_base_salary = flt(custom_base_salary)
		if custom_allowance_config:
			doc.custom_allowance_config = custom_allowance_config
	doc.insert()
	return {"name": doc.name, "employee_name": doc.employee_name}


@frappe.whitelist()
def update_employee(name: str, payload=None):
	"""Update an existing Employee record.

	Parameters
	----------
	name:
		Employee docname (e.g. ``"HR-EMP-00001"``).
	payload:
		Dict (or JSON string) of fields to update.  Only fields in the
		explicit whitelist are applied; arbitrary keys are silently ignored.
		Salary fields (``custom_base_salary``, ``custom_allowance_config``)
		are silently ignored unless the caller holds a payroll-visible role.

	Returns
	-------
	``{"name": doc.name}``
	"""
	if frappe.session.user == "Guest":
		frappe.throw(frappe._("Not permitted"), frappe.PermissionError)
	_assert_can_write("Employee", name, "write")
	if not name or not frappe.db.exists("Employee", name):
		frappe.throw(f"Unknown employee: {name}")

	# Normalise payload
	if payload is None:
		payload = {}
	if isinstance(payload, str):
		try:
			payload = json.loads(payload)
		except Exception:
			frappe.throw("payload must be a valid JSON object.")
	if not isinstance(payload, dict):
		frappe.throw("payload must be a JSON object.")

	can_see_salary = _user_can_see_salary()

	# Validate enum fields if present in payload
	shift_class = payload.get("custom_shift_class")
	if shift_class is not None and shift_class not in _VALID_SHIFT_CLASS:
		frappe.throw(f"Invalid custom_shift_class: {shift_class}. Must be one of {sorted(_VALID_SHIFT_CLASS)}.")

	region = payload.get("custom_region")
	if region is not None and region not in _VALID_REGION:
		frappe.throw(f"Invalid custom_region: {region}. Must be one of {sorted(_VALID_REGION)}.")

	work_mode = payload.get("custom_work_mode")
	if work_mode is not None and work_mode not in _VALID_WORK_MODE:
		frappe.throw(f"Invalid custom_work_mode: {work_mode}. Must be one of {sorted(_VALID_WORK_MODE)}.")

	# Validate stake_coefficient
	if "custom_stake_coefficient" in payload:
		# Determine effective work_mode: either from payload or from DB
		effective_work_mode = work_mode
		if effective_work_mode is None:
			effective_work_mode = frappe.db.get_value("Employee", name, "custom_work_mode")
		coeff = flt(payload["custom_stake_coefficient"])
		if effective_work_mode != "HALF_RATE":
			# Force 1.0 for non-HALF_RATE modes
			payload["custom_stake_coefficient"] = 1.0
		elif not (0.1 <= coeff <= 2.0):
			frappe.throw("custom_stake_coefficient must be between 0.1 and 2.0.")
	elif work_mode is not None and work_mode != "HALF_RATE":
		# Changing work_mode away from HALF_RATE resets coefficient to 1.0
		payload["custom_stake_coefficient"] = 1.0

	# Validate custom_allowance_config JSON
	if "custom_allowance_config" in payload and payload["custom_allowance_config"]:
		try:
			json.loads(payload["custom_allowance_config"])
		except Exception:
			frappe.throw("custom_allowance_config must be valid JSON.")

	# Apply whitelisted fields only
	doc = frappe.get_doc("Employee", name)
	for field, value in payload.items():
		if field not in _EMPLOYEE_WRITABLE_FIELDS:
			continue
		if field in SALARY_FIELDS and not can_see_salary:
			continue
		setattr(doc, field, value)

	doc.save()
	return {"name": doc.name}


@frappe.whitelist()
def list_designations(search: str = "", limit: int = 50):
	conds = []
	params: dict = {"limit": int(limit)}
	if search:
		conds.append("name LIKE %(s)s")
		params["s"] = f"%{search}%"
	where = (" WHERE " + " AND ".join(conds)) if conds else ""
	return frappe.db.sql(
		f"SELECT name FROM `tabDesignation`{where} ORDER BY name LIMIT %(limit)s",
		params,
		as_dict=True,
	)


@frappe.whitelist()
def list_departments(company: str = "", search: str = "", limit: int = 50):
	# Authn/authz: require the HR module (the rest of hr.py gates via
	# _require_company; this reader is company-optional so it needs its own
	# guard). Then scope by company: validate a passed company against the
	# caller's allowed set, and when omitted restrict a scoped non-admin to
	# their allowed companies (global / NULL-company departments stay visible).
	from stabler.api.organization import (
		_ADMIN_ROLES,
		_can_access_module,
		_user_allowed_companies,
	)

	if not _can_access_module(frappe.session.user, "hr"):
		frappe.throw(frappe._("Not permitted"), frappe.PermissionError)

	is_admin = any(r in frappe.get_roles() for r in _ADMIN_ROLES)
	allowed = [] if is_admin else _user_allowed_companies(frappe.session.user)

	conds = []
	params: dict = {"limit": int(limit)}
	if company:
		if allowed and company not in allowed:
			frappe.throw(
				frappe._("Not permitted for company {0}").format(company), frappe.PermissionError
			)
		conds.append("(company = %(c)s OR company IS NULL OR company = '')")
		params["c"] = company
	elif allowed:
		conds.append("(company IN %(allowed)s OR company IS NULL OR company = '')")
		params["allowed"] = tuple(allowed)
	if search:
		conds.append("name LIKE %(s)s")
		params["s"] = f"%{search}%"
	where = (" WHERE " + " AND ".join(conds)) if conds else ""
	return frappe.db.sql(
		f"SELECT name FROM `tabDepartment`{where} ORDER BY name LIMIT %(limit)s",
		params,
		as_dict=True,
	)


# ----- Attendance ----------------------------------------------------------


@frappe.whitelist()
def list_attendance(
	company: str,
	from_date: str = "",
	to_date: str = "",
	employee: str = "",
	status: str = "",
	limit: int = 200,
):
	_require_company(company)
	conds = ["company = %(company)s"]
	params: dict = {"company": company, "limit": int(limit)}
	if from_date:
		conds.append("attendance_date >= %(from_date)s")
		params["from_date"] = getdate(from_date)
	if to_date:
		conds.append("attendance_date <= %(to_date)s")
		params["to_date"] = getdate(to_date)
	if employee:
		conds.append("employee = %(employee)s")
		params["employee"] = employee
	if status:
		conds.append("status = %(status)s")
		params["status"] = status
	where = " AND ".join(conds)
	return frappe.db.sql(
		f"""
		SELECT name, employee, employee_name, attendance_date, status,
		       in_time, out_time, working_hours, leave_type, docstatus
		FROM `tabAttendance`
		WHERE {where}
		ORDER BY attendance_date DESC, employee
		LIMIT %(limit)s
		""",
		params,
		as_dict=True,
	)


_ATT_CODE = {
	"Present": "P",
	"Absent": "A",
	"On Leave": "L",
	"Half Day": "H",
	"Work From Home": "W",
}


@frappe.whitelist()
def attendance_matrix(company: str, period: str = "", department: str = "", employee: str = "") -> dict:
	"""Whole-roster month grid: one row per employee, one column per day.

	Returns day metadata (weekend/holiday/today), per-employee day codes
	(P/A/L/H/W), late-entry days, and per-employee totals. Company-scoped.
	"""
	import calendar as _cal

	_require_company(company)
	if not period:
		period = today()[:7]
	year, mon = (int(x) for x in period.split("-")[:2])
	days_in_month = _cal.monthrange(year, mon)[1]
	from_date = getdate(f"{year:04d}-{mon:02d}-01")
	to_date = getdate(f"{year:04d}-{mon:02d}-{days_in_month:02d}")
	today_d = getdate(today())

	# Holiday list (weekly-off vs real holiday) for shading.
	weekly_off_days: set[int] = set()
	holiday_days: set[int] = set()
	holiday_list = frappe.get_cached_value("Company", company, "default_holiday_list")
	if holiday_list:
		for h in frappe.get_all(
			"Holiday",
			filters={"parent": holiday_list, "holiday_date": ["between", [from_date, to_date]]},
			fields=["holiday_date", "weekly_off"],
		):
			d = getdate(h.holiday_date).day
			(weekly_off_days if h.weekly_off else holiday_days).add(d)

	lock_date = _attendance_lock_date()
	days = []
	for d in range(1, days_in_month + 1):
		dt = getdate(f"{year:04d}-{mon:02d}-{d:02d}")
		wd = dt.weekday()  # 0=Mon … 6=Sun
		is_weekend = d in weekly_off_days or wd >= 5
		days.append(
			{
				"day": d,
				"date": str(dt),
				"weekday": "MTWTFSS"[wd],
				"is_weekend": bool(is_weekend),
				"is_holiday": d in holiday_days,
				"is_today": dt == today_d,
				"is_future": dt > today_d,
				"is_locked": bool(lock_date and dt <= lock_date),
			}
		)

	# Roster (active employees), optionally filtered by department.
	emp_filters: dict = {"company": company, "status": "Active"}
	if department:
		emp_filters["department"] = department
	if employee:
		emp_filters["name"] = employee  # single-employee calendar view
	roster = frappe.get_all(
		"Employee",
		filters=emp_filters,
		fields=["name", "employee_name", "department", "designation"],
		order_by="department asc, employee_name asc",
		limit_page_length=0,
	)

	# Attendance for the month.
	records = frappe.db.sql(
		"""
		SELECT employee, attendance_date, status, late_entry, working_hours, leave_type
		FROM `tabAttendance`
		WHERE company = %(company)s AND docstatus < 2
		  AND attendance_date BETWEEN %(from_date)s AND %(to_date)s
		""",
		{"company": company, "from_date": from_date, "to_date": to_date},
		as_dict=True,
	)
	by_emp: dict[str, dict] = {}
	for r in records:
		day = getdate(r.attendance_date).day
		by_emp.setdefault(r.employee, {})[day] = r

	rows = []
	for emp in roster:
		cells: dict[int, str] = {}
		late: list[int] = []
		totals = {"present": 0, "absent": 0, "leave": 0, "half": 0, "wfh": 0, "late": 0}
		emp_recs = by_emp.get(emp.name, {})
		for day, rec in emp_recs.items():
			code = _ATT_CODE.get(rec.status, "")
			if not code:
				continue
			cells[day] = code
			if code == "P":
				totals["present"] += 1
			elif code == "A":
				totals["absent"] += 1
			elif code == "L":
				totals["leave"] += 1
			elif code == "H":
				totals["half"] += 1
			elif code == "W":
				totals["wfh"] += 1
				totals["present"] += 1
			if rec.late_entry:
				late.append(day)
				totals["late"] += 1
		rows.append(
			{
				"employee": emp.name,
				"employee_name": emp.employee_name,
				"department": (emp.department or "").split(" - ")[0] if emp.department else "",
				"designation": (emp.designation or "").split(" - ")[0] if emp.designation else "",
				"cells": cells,
				"late": late,
				"totals": totals,
			}
		)

	return {
		"company": company,
		"period": f"{year:04d}-{mon:02d}",
		"days_in_month": days_in_month,
		"days": days,
		"rows": rows,
		"edit_lock_date": str(lock_date) if lock_date else None,
		"today": str(today_d),
	}


@frappe.whitelist()
def mark_attendance(
	company: str,
	employee: str,
	attendance_date: str,
	status: str,
	in_time: str = "",
	out_time: str = "",
	submit: int = 1,
):
	_require_company(company)
	if not employee or not frappe.db.exists("Employee", employee):
		frappe.throw(f"Unknown employee: {employee}")
	if status not in ("Present", "Absent", "On Leave", "Half Day", "Work From Home"):
		frappe.throw(f"Invalid status: {status}")
	att_date = getdate(attendance_date or today())
	existing = frappe.db.get_value(
		"Attendance",
		{"employee": employee, "attendance_date": att_date, "docstatus": ("<", 2)},
		"name",
	)
	if existing:
		frappe.throw(f"Attendance already recorded for {employee} on {att_date} ({existing}).")

	doc = frappe.new_doc("Attendance")
	doc.company = company
	doc.employee = employee
	doc.attendance_date = att_date
	doc.status = status
	if in_time:
		doc.in_time = in_time
	if out_time:
		doc.out_time = out_time
	doc.insert()
	if int(submit or 0):
		doc.submit()
	return {"name": doc.name, "status": doc.status, "docstatus": doc.docstatus}


_VALID_ATT_STATUS = ("Present", "Absent", "On Leave", "Half Day", "Work From Home")


def _attendance_lock_date():
	"""On/before this date attendance is frozen for inline edits. Sourced from the
	optional `attendance_edit_lock_days` Stabler Settings field; None = no hard lock
	(only the front-end confirmation applies)."""
	try:
		days = frappe.db.get_single_value("Stabler Settings", "attendance_edit_lock_days")
	except Exception:
		days = None
	if not days or int(days) <= 0:
		return None
	return getdate(add_days(today(), -int(days)))


def _assert_attendance_editable(att_date) -> None:
	if att_date > getdate(today()):
		frappe.throw("Cannot edit attendance for a future date.")
	lock = _attendance_lock_date()
	if lock and att_date <= lock:
		frappe.throw(f"Attendance on/before {lock} is locked and cannot be edited.")


def _drop_existing_attendance(employee: str, att_date) -> None:
	existing = frappe.db.get_value(
		"Attendance",
		{"employee": employee, "attendance_date": att_date, "docstatus": ("<", 2)},
		["name", "docstatus"],
		as_dict=True,
	)
	if not existing:
		return
	doc = frappe.get_doc("Attendance", existing.name)
	if existing.docstatus == 1:
		doc.cancel()
	doc.delete(ignore_permissions=False)


_ATT_MIN_FIELDS = ("custom_late_minutes", "custom_early_minutes", "custom_overtime_minutes")


def _hm(dt) -> str:
	return dt.strftime("%H:%M") if dt else ""


@frappe.whitelist()
def set_attendance(
	company: str,
	employee: str,
	attendance_date: str,
	status: str,
	in_time: str = "",
	out_time: str = "",
	late_minutes=0,
	early_minutes=0,
	overtime_minutes=0,
) -> dict:
	"""Upsert one day's attendance (inline cell edit on the month grid).

	Records status plus arrival/departure times and the late / early-leave /
	overtime minutes. Replaces any existing record for that employee+date, then
	submits — exactly one authoritative record per day. late_entry/early_exit are
	set from the minute inputs so a day can be flagged late.
	"""
	_require_company(company)
	if not employee or not frappe.db.exists("Employee", employee):
		frappe.throw(f"Unknown employee: {employee}")
	if status not in _VALID_ATT_STATUS:
		frappe.throw(f"Invalid status: {status}")
	if not frappe.has_permission("Attendance", "create"):
		frappe.throw("Not permitted to edit attendance.", frappe.PermissionError)
	att_date = getdate(attendance_date or today())
	_assert_attendance_editable(att_date)

	lm = max(0, int(late_minutes or 0))
	em = max(0, int(early_minutes or 0))
	ot = max(0, int(overtime_minutes or 0))

	_drop_existing_attendance(employee, att_date)
	doc = frappe.new_doc("Attendance")
	doc.company = company
	doc.employee = employee
	doc.attendance_date = att_date
	doc.status = status
	if in_time:
		doc.in_time = f"{att_date} {in_time}:00"
	if out_time:
		doc.out_time = f"{att_date} {out_time}:00"
	doc.late_entry = 1 if lm > 0 else 0
	doc.early_exit = 1 if em > 0 else 0
	for fn, val in (("custom_late_minutes", lm), ("custom_early_minutes", em), ("custom_overtime_minutes", ot)):
		if frappe.db.has_column("Attendance", fn):
			doc.set(fn, val)
	doc.insert()
	doc.submit()
	return {
		"name": doc.name, "employee": employee, "day": att_date.day,
		"status": status, "late": lm > 0, "docstatus": doc.docstatus,
	}


@frappe.whitelist()
def get_attendance_cell(company: str, employee: str, attendance_date: str) -> dict:
	"""Prefill data for the inline cell editor: one day's record + lock state."""
	_require_company(company)
	att_date = getdate(attendance_date)
	lock = _attendance_lock_date()
	locked = bool(lock and att_date <= lock) or att_date > getdate(today())
	cols = ["status", "in_time", "out_time", "late_entry", "early_exit"]
	cols += [f for f in _ATT_MIN_FIELDS if frappe.db.has_column("Attendance", f)]
	rec = frappe.db.get_value(
		"Attendance",
		{"employee": employee, "attendance_date": att_date, "docstatus": ("<", 2)},
		cols,
		as_dict=True,
	)
	if not rec:
		return {"status": None, "in_time": "", "out_time": "", "late_minutes": 0,
			"early_minutes": 0, "overtime_minutes": 0, "locked": locked}
	return {
		"status": rec.status,
		"in_time": _hm(rec.in_time),
		"out_time": _hm(rec.out_time),
		"late_minutes": int(rec.get("custom_late_minutes") or 0),
		"early_minutes": int(rec.get("custom_early_minutes") or 0),
		"overtime_minutes": int(rec.get("custom_overtime_minutes") or 0),
		"locked": locked,
	}


@frappe.whitelist()
def bulk_set_attendance(company: str, items, status: str = "") -> dict:
	"""Mark (or clear, when status is empty) many employee+day cells at once.

	Status-only — used by the matrix multi-select. Each cell respects the future +
	lock guards; non-editable or unknown cells are reported, the rest are written.
	"""
	_require_company(company)
	if isinstance(items, str):
		items = frappe.parse_json(items) or []
	if not isinstance(items, list) or not items:
		frappe.throw("No cells selected.")
	if len(items) > 3000:
		frappe.throw("Too many cells (max 3000).")
	clearing = not status
	if not clearing and status not in _VALID_ATT_STATUS:
		frappe.throw(f"Invalid status: {status}")
	if not frappe.has_permission("Attendance", "delete" if clearing else "create"):
		frappe.throw("Not permitted to edit attendance.", frappe.PermissionError)

	updated = 0
	skipped = []
	for it in items:
		emp = (it or {}).get("employee")
		d = (it or {}).get("date")
		if not emp or not d:
			skipped.append({"employee": emp, "date": d, "reason": "incomplete cell"})
			continue
		att_date = getdate(d)
		try:
			_assert_attendance_editable(att_date)
		except Exception as exc:
			skipped.append({"employee": emp, "date": d, "reason": str(exc)})
			continue
		if not frappe.db.exists("Employee", emp):
			skipped.append({"employee": emp, "date": d, "reason": "unknown employee"})
			continue
		_drop_existing_attendance(emp, att_date)
		if not clearing:
			doc = frappe.new_doc("Attendance")
			doc.company = company
			doc.employee = emp
			doc.attendance_date = att_date
			doc.status = status
			doc.insert()
			doc.submit()
		updated += 1
	frappe.db.commit()
	return {"updated": updated, "skipped": skipped, "status": status or None}


@frappe.whitelist()
def clear_attendance(company: str, employee: str, attendance_date: str) -> dict:
	"""Remove a day's attendance record (inline 'clear' on the grid)."""
	_require_company(company)
	if not frappe.has_permission("Attendance", "delete"):
		frappe.throw("Not permitted to edit attendance.", frappe.PermissionError)
	att_date = getdate(attendance_date or today())
	_assert_attendance_editable(att_date)
	_drop_existing_attendance(employee, att_date)
	return {"employee": employee, "day": att_date.day, "cleared": True}


_ATT_XLSX_STYLE = {
	"P": ("E7F6EC", "1F9D54"),
	"A": ("FBE9E9", "D63939"),
	"L": ("E7EDFB", "3B5BDB"),
	"H": ("FDF3E3", "C07A00"),
	"W": ("E6F7F4", "0CA678"),
}

# Display letter shown in the sheet — follows the anjan-hr standard:
#   K = keldi (present / WFH / late), Y = yarim kun (half day),
#   D = kelmadi (absent / on leave).  The P/A/L/H/W code stays the data key.
_ATT_XLSX_LETTER = {"P": "K", "W": "K", "H": "Y", "A": "D", "L": "D"}


@frappe.whitelist()
def attendance_matrix_xlsx(company: str, period: str = "", department: str = ""):
	"""Download the month attendance grid as a colour-coded .xlsx."""
	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
	from openpyxl.utils import get_column_letter

	m = attendance_matrix(company, period, department)
	days = m["days"]
	rows = m["rows"]
	ndays = len(days)

	wb = Workbook()
	ws = wb.active
	ws.title = "Attendance"

	thin = Side(style="thin", color="E6E7E9")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)
	center = Alignment(horizontal="center", vertical="center")
	weekend_fill = PatternFill("solid", fgColor="F3F5F8")
	header_fill = PatternFill("solid", fgColor="F7F9FC")

	# Row 1 — title.
	ws.cell(row=1, column=1, value=f"{company} — Attendance {m['period']}").font = Font(bold=True, size=13)

	FIXED = ["Employee", "Department", "P", "A", "L"]
	# Row 2 — headers + day numbers; Row 3 — weekday letters.
	for i, h in enumerate(FIXED, start=1):
		c = ws.cell(row=2, column=i, value=h)
		c.font = Font(bold=True)
		c.fill = header_fill
		c.alignment = center
		c.border = border
	for di, d in enumerate(days):
		col = 6 + di
		c = ws.cell(row=2, column=col, value=d["day"])
		c.font = Font(bold=True, size=9)
		c.alignment = center
		c.border = border
		c.fill = weekend_fill if d["is_weekend"] else header_fill
		w = ws.cell(row=3, column=col, value=d["weekday"])
		w.font = Font(size=8, color="9AA4B2")
		w.alignment = center
		w.border = border
		w.fill = weekend_fill if d["is_weekend"] else header_fill
	for i in range(1, 6):
		ws.cell(row=3, column=i).fill = header_fill
		ws.cell(row=3, column=i).border = border

	# Data rows.
	r = 4
	for row in rows:
		ws.cell(row=r, column=1, value=row["employee_name"]).border = border
		ws.cell(row=r, column=2, value=row.get("department") or "").border = border
		for col, key, color in ((3, "present", "1F9D54"), (4, "absent", "D63939"), (5, "leave", "3B5BDB")):
			cc = ws.cell(row=r, column=col, value=row["totals"].get(key) or None)
			cc.alignment = center
			cc.font = Font(bold=True, color=color)
			cc.border = border
		late = set(row.get("late") or [])
		for di, d in enumerate(days):
			col = 6 + di
			code = (row.get("cells") or {}).get(str(d["day"])) or (row.get("cells") or {}).get(d["day"])
			cell = ws.cell(row=r, column=col)
			cell.alignment = center
			cell.border = border
			if code in _ATT_XLSX_STYLE:
				bg, fg = _ATT_XLSX_STYLE[code]
				letter = _ATT_XLSX_LETTER.get(code, code)
				cell.value = letter
				cell.fill = PatternFill("solid", fgColor=bg)
				cell.font = Font(bold=True, color=fg, size=9)
				if d["day"] in late:
					cell.value = f"{letter}*"
			elif d["is_weekend"]:
				cell.fill = weekend_fill
		r += 1

	# Legend.
	lr = r + 1
	ws.cell(row=lr, column=1, value="Legend:").font = Font(bold=True)
	leg = [("P", "Present"), ("A", "Absent"), ("L", "On Leave"), ("H", "Half Day"), ("W", "WFH"), ("*", "Late")]
	for i, (code, label) in enumerate(leg):
		letter = _ATT_XLSX_LETTER.get(code, code)
		c = ws.cell(row=lr, column=2 + i, value=f"{letter} {label}")
		if code in _ATT_XLSX_STYLE:
			bg, fg = _ATT_XLSX_STYLE[code]
			c.fill = PatternFill("solid", fgColor=bg)
			c.font = Font(color=fg, size=9)

	# Layout: widths + frozen panes.
	ws.column_dimensions["A"].width = 24
	ws.column_dimensions["B"].width = 16
	for i in (3, 4, 5):
		ws.column_dimensions[get_column_letter(i)].width = 4
	for di in range(ndays):
		ws.column_dimensions[get_column_letter(6 + di)].width = 3.6
	ws.freeze_panes = "F4"

	import io

	buf = io.BytesIO()
	wb.save(buf)
	frappe.local.response.filename = f"attendance-{m['period']}.xlsx"
	frappe.local.response.filecontent = buf.getvalue()
	frappe.local.response.type = "binary"


# ----- Leave Applications --------------------------------------------------


@frappe.whitelist()
def list_leave_applications(
	company: str,
	status: str = "",
	employee: str = "",
	from_date: str = "",
	to_date: str = "",
	limit: int = 100,
):
	_require_company(company)
	conds = ["company = %(company)s"]
	params: dict = {"company": company, "limit": int(limit)}
	if status:
		conds.append("status = %(status)s")
		params["status"] = status
	if employee:
		conds.append("employee = %(employee)s")
		params["employee"] = employee
	if from_date:
		conds.append("from_date >= %(from_date)s")
		params["from_date"] = getdate(from_date)
	if to_date:
		conds.append("to_date <= %(to_date)s")
		params["to_date"] = getdate(to_date)
	where = " AND ".join(conds)
	return frappe.db.sql(
		f"""
		SELECT name, employee, employee_name, leave_type, from_date, to_date,
		       total_leave_days, status, posting_date, docstatus, description
		FROM `tabLeave Application`
		WHERE {where}
		ORDER BY posting_date DESC, name DESC
		LIMIT %(limit)s
		""",
		params,
		as_dict=True,
	)


@frappe.whitelist()
def list_leave_types(limit: int = 50):
	return frappe.db.sql(
		"SELECT name, max_continuous_days_allowed, is_lwp, allow_negative FROM `tabLeave Type` ORDER BY name LIMIT %(limit)s",
		{"limit": int(limit)},
		as_dict=True,
	)


@frappe.whitelist()
def create_leave_application(
	company: str,
	employee: str,
	leave_type: str,
	from_date: str,
	to_date: str,
	description: str = "",
	submit: int = 0,
):
	_require_company(company)
	if not employee or not frappe.db.exists("Employee", employee):
		frappe.throw(f"Unknown employee: {employee}")
	if not leave_type or not frappe.db.exists("Leave Type", leave_type):
		frappe.throw(f"Unknown leave type: {leave_type}")
	fd, td = getdate(from_date), getdate(to_date)
	if td < fd:
		frappe.throw("To date cannot be before from date.")

	doc = frappe.new_doc("Leave Application")
	doc.company = company
	doc.employee = employee
	doc.leave_type = leave_type
	doc.from_date = fd
	doc.to_date = td
	doc.posting_date = today()
	doc.status = "Open"
	if description:
		doc.description = description
	doc.insert()
	if int(submit or 0):
		doc.status = "Approved"
		doc.submit()
	return {"name": doc.name, "status": doc.status, "docstatus": doc.docstatus}


@frappe.whitelist()
def approve_leave(name: str, status: str = "Approved"):
	_assert_can_write("Leave Application", name, "submit")
	if not name or not frappe.db.exists("Leave Application", name):
		frappe.throw(f"Unknown leave application: {name}")
	if status not in ("Approved", "Rejected"):
		frappe.throw(f"Invalid status: {status}")
	doc = frappe.get_doc("Leave Application", name)
	if doc.docstatus == 1:
		frappe.throw("Already submitted.")
	doc.status = status
	if status == "Approved":
		doc.submit()
	else:
		doc.save()
	return {"name": doc.name, "status": doc.status, "docstatus": doc.docstatus}


# ----- Payroll -------------------------------------------------------------


@frappe.whitelist()
def list_salary_slips(
	company: str,
	from_date: str = "",
	to_date: str = "",
	employee: str = "",
	status: str = "",
	limit: int = 100,
):
	_require_company(company)
	conds = ["company = %(company)s"]
	params: dict = {"company": company, "limit": int(limit)}
	if from_date:
		conds.append("start_date >= %(from_date)s")
		params["from_date"] = getdate(from_date)
	if to_date:
		conds.append("end_date <= %(to_date)s")
		params["to_date"] = getdate(to_date)
	if employee:
		conds.append("employee = %(employee)s")
		params["employee"] = employee
	if status:
		conds.append("status = %(status)s")
		params["status"] = status
	where = " AND ".join(conds)
	return frappe.db.sql(
		f"""
		SELECT name, employee, employee_name, posting_date, start_date, end_date,
		       total_working_days, payment_days, gross_pay, net_pay, currency,
		       status, docstatus
		FROM `tabSalary Slip`
		WHERE {where}
		ORDER BY posting_date DESC, name DESC
		LIMIT %(limit)s
		""",
		params,
		as_dict=True,
	)


@frappe.whitelist()
def salary_slip_detail(name: str):
	_assert_can_read("Salary Slip", name)
	if not name or not frappe.db.exists("Salary Slip", name):
		frappe.throw(f"Unknown salary slip: {name}")
	doc = frappe.get_doc("Salary Slip", name)
	earnings = [
		{"component": r.salary_component, "amount": flt(r.amount)}
		for r in (doc.earnings or [])
	]
	deductions = [
		{"component": r.salary_component, "amount": flt(r.amount)}
		for r in (doc.deductions or [])
	]
	return {
		"name": doc.name,
		"employee": doc.employee,
		"employee_name": doc.employee_name,
		"posting_date": doc.posting_date,
		"start_date": doc.start_date,
		"end_date": doc.end_date,
		"company": doc.company,
		"currency": doc.currency,
		"total_working_days": flt(doc.total_working_days),
		"payment_days": flt(doc.payment_days),
		"gross_pay": flt(doc.gross_pay),
		"total_deduction": flt(doc.total_deduction),
		"net_pay": flt(doc.net_pay),
		"status": doc.status,
		"docstatus": doc.docstatus,
		"earnings": earnings,
		"deductions": deductions,
	}


@frappe.whitelist()
def create_payroll_entry(
	company: str,
	start_date: str,
	end_date: str,
	posting_date: str = "",
	payroll_frequency: str = "Monthly",
	submit: int = 0,
):
	"""Create a Payroll Entry covering employees with salary structures.

	Returns the payroll entry name plus generated salary slip names.
	"""
	_require_company(company)
	fd, td = getdate(start_date), getdate(end_date)
	if td < fd:
		frappe.throw("End date cannot be before start date.")

	doc = frappe.new_doc("Payroll Entry")
	doc.company = company
	doc.start_date = fd
	doc.end_date = td
	doc.posting_date = getdate(posting_date or today())
	doc.payroll_frequency = payroll_frequency

	# Frappe will pull eligible employees via the controller
	if hasattr(doc, "fill_employee_details"):
		doc.fill_employee_details()
	doc.insert()

	created_slips: list[str] = []
	try:
		if hasattr(doc, "create_salary_slips"):
			doc.create_salary_slips()
		created_slips = [
			s.name
			for s in frappe.get_all(
				"Salary Slip",
				filters={"payroll_entry": doc.name},
				pluck="name",
			)
		]
		if int(submit or 0) and hasattr(doc, "submit_salary_slips"):
			doc.submit_salary_slips()
	except Exception as exc:
		frappe.log_error(frappe.get_traceback(), "Payroll Entry generation failed")
		return {
			"name": doc.name,
			"warning": f"Created payroll entry but slip generation hit: {exc}",
			"salary_slips": created_slips,
		}

	return {"name": doc.name, "salary_slips": created_slips}


# ----- Dashboard helpers ---------------------------------------------------


@frappe.whitelist()
def hr_overview(company: str):
	"""Quick stats for the HR home tab header."""
	_require_company(company)
	active = frappe.db.count("Employee", {"company": company, "status": "Active"})
	on_leave_today = frappe.db.sql(
		"""
		SELECT COUNT(DISTINCT employee) FROM `tabLeave Application`
		WHERE company = %s AND status = 'Approved' AND docstatus = 1
		  AND from_date <= %s AND to_date >= %s
		""",
		(company, today(), today()),
	)[0][0]
	pending_leave = frappe.db.count(
		"Leave Application", {"company": company, "status": "Open", "docstatus": 0}
	)
	return {
		"active_employees": int(active or 0),
		"on_leave_today": int(on_leave_today or 0),
		"pending_leave_requests": int(pending_leave or 0),
	}


@frappe.whitelist()
def import_timepay_names(rows, apply=0, apply_name_matches=0):
	"""Backfill TimePay names onto Employees from an anjan-hr export.

	`rows`: JSON string (or list) of {timepay_id, fio} exported from anjan-hr.
	`apply=0`  → dry run: returns the plan without writing.
	`apply=1`  → writes the SAFE id-matched name updates
	             (stabler.custom_timepay_id == anjan.timepay_id → set custom_timepay_name).
	`apply_name_matches=1` → ALSO apply the name-fallback suggestions (sets both
	             custom_timepay_id and custom_timepay_name). Review these first.

	Admin / HR Manager only.
	"""
	from stabler.api.organization import _ADMIN_ROLES
	from stabler.api._timepay_import import plan_timepay_import

	if not (set(frappe.get_roles()) & ({"HR Manager"} | set(_ADMIN_ROLES))):
		frappe.throw("Not permitted — admin or HR Manager only.", frappe.PermissionError)

	if isinstance(rows, str):
		try:
			rows = json.loads(rows)
		except Exception:
			frappe.throw("rows must be valid JSON (a list of {timepay_id, fio}).")
	rows = list(rows or [])

	emps = frappe.get_all(
		"Employee",
		fields=["name", "employee_name", "custom_timepay_id", "custom_timepay_name"],
		limit_page_length=0,
	)
	plan = plan_timepay_import(emps, rows)

	applied_names = 0
	applied_id_name = 0
	if int(apply or 0):
		for u in plan["id_updates"]:
			frappe.db.set_value("Employee", u["employee"], "custom_timepay_name", u["new_name"])
			applied_names += 1
		if int(apply_name_matches or 0):
			for s in plan["name_suggestions"]:
				vals = {"custom_timepay_name": s["fio"]}
				if s.get("timepay_id"):
					vals["custom_timepay_id"] = s["timepay_id"]
				frappe.db.set_value("Employee", s["employee"], vals)
				applied_id_name += 1
		frappe.db.commit()

	return {
		"applied": bool(int(apply or 0)),
		"counts": {
			"anjan_rows": len(rows),
			"id_updates": len(plan["id_updates"]),
			"name_suggestions": len(plan["name_suggestions"]),
			"unmatched": len(plan["unmatched"]),
			"applied_names": applied_names,
			"applied_name_matches": applied_id_name,
		},
		"plan": plan,
	}
