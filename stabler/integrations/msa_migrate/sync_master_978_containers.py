"""Sync all 978 Master Containers from 'jas' tab of 'Оплаты заводам HMA, FAIR, MIRHA, IFF _ PI (1).xlsx' to Import Container records on msa.erpstable.com."""

import json
import os

import frappe
from frappe.utils import cint, flt, today

SUPPLIER_MAP = {
	"HMA AGRO INDUSTRIES LIMITED": "HMA AGRO INDUSTRIES LIMITED",
	"Mirha Exports Private limited": "Mirha Exports Private limited",
	"FAIR EXPORTS (INDIA) PVT. LTD.": "FAIR EXPORTS (INDIA) PVT. LTD.",
	"IFF India Frozen Foods Private Limited": "IFF India Frozen Foods Private Limited",
	"AL-SUPER FROZEN FOODS PVT. LTD": "Al Super Frozen Food Private Limited",
	"AL-DUA": "AL DUA FOOD PROCESSING PVT. LTD",
}


def get_master_containers():
	json_file = "/tmp/master_978_containers.json"
	if os.path.exists(json_file):
		with open(json_file, encoding="utf-8") as f:
			return json.load(f)

	# Fallback inline reader
	import openpyxl

	excel_path = "/Users/zafar/Downloads/Оплаты заводам HMA, FAIR, MIRHA, IFF _ PI (1).xlsx"
	wb = openpyxl.load_workbook(excel_path, data_only=True)
	ws = wb["jas"]

	containers_map = {}
	for r in range(3, ws.max_row + 1):
		contract = str(ws.cell(row=r, column=2).value or "").strip()
		pi_no = str(ws.cell(row=r, column=3).value or "").strip()
		ci_no = str(ws.cell(row=r, column=4).value or "").strip()
		cdate = ws.cell(row=r, column=5).value
		cnt_no = str(ws.cell(row=r, column=6).value or "").strip()
		supplier = str(ws.cell(row=r, column=7).value or "").strip()
		item_name = str(ws.cell(row=r, column=8).value or "").strip()
		box_qty = ws.cell(row=r, column=12).value
		total_kg = ws.cell(row=r, column=14).value
		rate = ws.cell(row=r, column=15).value
		amount = ws.cell(row=r, column=16).value

		if contract == "None":
			contract = ""
		if pi_no == "None":
			pi_no = ""
		if ci_no == "None":
			ci_no = ""
		if cnt_no == "None":
			cnt_no = ""
		if supplier == "None":
			supplier = ""
		if item_name == "None":
			item_name = ""

		if ci_no and len(ci_no) > 7 and "-" not in ci_no[-7:]:
			if ci_no[-6:].isdigit() and ci_no[-6:-2] == "2025" and ci_no[-2:] in ["26", "27"]:
				ci_no = ci_no[:-6] + ci_no[-6:-2] + "-" + ci_no[-2:]

		if cnt_no:
			if cnt_no not in containers_map:
				containers_map[cnt_no] = {
					"container_number": cnt_no,
					"commercial_invoice": ci_no,
					"proforma_invoice": pi_no,
					"supplier": supplier,
					"contract": contract,
					"date": str(cdate)[:10] if cdate else "",
					"items": [],
				}
			try:
				b_qty = int(box_qty) if box_qty else 0
			except (TypeError, ValueError):
				b_qty = 0
			try:
				t_kg = float(total_kg) if total_kg else 0.0
			except (TypeError, ValueError):
				t_kg = 0.0
			try:
				r_rate = float(rate) if rate else 0.0
			except (TypeError, ValueError):
				r_rate = 0.0
			try:
				a_amt = float(amount) if amount else 0.0
			except (TypeError, ValueError):
				a_amt = 0.0

			if item_name or b_qty > 0 or t_kg > 0:
				containers_map[cnt_no]["items"].append(
					{
						"item_name": item_name,
						"box_qty": b_qty,
						"total_kg": t_kg,
						"rate": r_rate,
						"amount": a_amt,
					}
				)
	return list(containers_map.values())


def ensure_item_exists(item_name: str) -> str:
	if not item_name:
		item_name = "BUFFALO MEAT"
	item_code = item_name.strip()
	if frappe.db.exists("Item", item_code):
		return item_code
	found = frappe.get_all("Item", filters={"item_name": item_name}, fields=["name"])
	if found:
		return found[0]["name"]
	# Create missing Item
	doc = frappe.get_doc(
		{
			"doctype": "Item",
			"item_code": item_code,
			"item_name": item_name,
			"item_group": "Meat Products",
			"stock_uom": "Kg",
			"is_stock_item": 1,
		}
	)
	doc.insert(ignore_permissions=True)
	return doc.name


def run(dry_run=1):
	dry_run = int(dry_run)
	company = "MSA"

	containers = get_master_containers()

	updated_count = 0
	created_count = 0

	for c in containers:
		cnt_no = c["container_number"]
		ci_raw = c.get("commercial_invoice") or ""
		supp_raw = c.get("supplier") or ""
		items = c.get("items") or []

		supp_name = SUPPLIER_MAP.get(supp_raw, supp_raw)

		# Resolve Commercial Invoice Doc
		ci_doc_name = None
		if ci_raw:
			if frappe.db.exists("Commercial Invoice", ci_raw):
				ci_doc_name = ci_raw
			else:
				found_ci = frappe.get_all(
					"Commercial Invoice", filters={"ci_number": ci_raw}, fields=["name"]
				)
				if found_ci:
					ci_doc_name = found_ci[0]["name"]

		# Calculate Totals
		tot_boxes = sum(cint(it["box_qty"]) for it in items)
		tot_kg = sum(flt(it["total_kg"]) for it in items)
		tot_amt = sum(flt(it["amount"]) for it in items)

		existing = frappe.get_all("Import Container", filters={"container_number": cnt_no}, fields=["name"])

		# Prepare item rows with valid item_code
		item_rows = []
		for it in items:
			raw_iname = it.get("item_name") or "BUFFALO MEAT"
			valid_icode = raw_iname if dry_run else ensure_item_exists(raw_iname)
			item_rows.append(
				{
					"item_code": valid_icode,
					"item_name": raw_iname,
					"box_qty": cint(it.get("box_qty")),
					"total_kg": flt(it.get("total_kg")),
					"rate": flt(it.get("rate")),
					"amount": flt(it.get("amount")),
				}
			)

		if existing:
			c_name = existing[0]["name"]
			updated_count += 1
			if not dry_run:
				updates = {
					"company": company,
					"supplier": supp_name if supp_name and frappe.db.exists("Supplier", supp_name) else None,
					"commercial_invoice": ci_doc_name,
					"total_boxes": tot_boxes,
					"total_kg": tot_kg,
					"total_amount": tot_amt,
					"currency": "USD",
				}
				frappe.db.set_value("Import Container", c_name, updates, update_modified=True)

				# Re-create item child table
				doc = frappe.get_doc("Import Container", c_name)
				doc.items = []
				for r_it in item_rows:
					doc.append("items", r_it)
				doc.flags.ignore_validate = True
				doc.save(ignore_permissions=True)
		else:
			created_count += 1
			if not dry_run:
				doc = frappe.get_doc(
					{
						"doctype": "Import Container",
						"container_number": cnt_no,
						"company": company,
						"supplier": supp_name
						if supp_name and frappe.db.exists("Supplier", supp_name)
						else None,
						"commercial_invoice": ci_doc_name,
						"status": "IN_TRANSIT",
						"total_boxes": tot_boxes,
						"total_kg": tot_kg,
						"total_amount": tot_amt,
						"currency": "USD",
						"items": item_rows,
					}
				)
				doc.flags.ignore_validate = True
				doc.insert(ignore_permissions=True)

	if not dry_run:
		frappe.db.commit()

	mode = "DRY-RUN" if dry_run else "APPLIED"

	print("\n========================================================")
	print(f"  SYNC ALL 978 MASTER CONTAINERS ({mode})")
	print("========================================================")
	print(f"Total Master Containers: {len(containers)}")
	print(f"Updated Containers: {updated_count}")
	print(f"Created Containers: {created_count}\n")

	return {
		"total_master": len(containers),
		"updated_count": updated_count,
		"created_count": created_count,
	}
