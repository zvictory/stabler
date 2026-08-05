"""Clean up non-master containers so total count equals exactly 978, and bind CI ↔ Container ↔ Transporter ↔ Freight Cost."""

import json
import os
import frappe
from frappe.utils import flt, cint

SUPPLIER_MAP = {
	"HMA AGRO INDUSTRIES LIMITED": "HMA AGRO INDUSTRIES LIMITED",
	"Mirha Exports Private limited": "Mirha Exports Private limited",
	"FAIR EXPORTS (INDIA) PVT. LTD.": "FAIR EXPORTS (INDIA) PVT. LTD.",
	"IFF India Frozen Foods Private Limited": "IFF India Frozen Foods Private Limited",
	"AL-SUPER FROZEN FOODS PVT. LTD": "Al Super Frozen Food Private Limited",
	"AL-DUA": "AL DUA FOOD PROCESSING PVT. LTD",
}


def get_master_container_numbers():
	json_file = "/tmp/master_978_containers.json"
	if os.path.exists(json_file):
		with open(json_file, "r", encoding="utf-8") as f:
			data = json.load(f)
			return [c["container_number"] for c in data if c.get("container_number")]

	# Fallback inline reader
	import openpyxl
	excel_path = "/Users/zafar/Downloads/Оплаты заводам HMA, FAIR, MIRHA, IFF _ PI (1).xlsx"
	wb = openpyxl.load_workbook(excel_path, data_only=True)
	ws = wb["jas"]
	cset = set()
	for r in range(3, ws.max_row + 1):
		cnt_no = str(ws.cell(row=r, column=6).value or "").strip()
		if cnt_no and cnt_no != "None":
			cset.add(cnt_no)
	return list(cset)


def run(dry_run=1):
	dry_run = int(dry_run)
	company = "MSA"

	master_numbers = get_master_container_numbers()
	master_set = set(master_numbers)

	print(f"Master 978 Container Numbers Count: {len(master_set)}")

	# 1. Find non-master containers to delete
	all_db_containers = frappe.get_all("Import Container", fields=["name", "container_number"])
	non_master_to_delete = []

	for c in all_db_containers:
		cnt_no = c["container_number"]
		if cnt_no not in master_set:
			non_master_to_delete.append(c["name"])

	deleted_count = len(non_master_to_delete)
	if not dry_run and non_master_to_delete:
		for cname in non_master_to_delete:
			# Unlink freight bookings
			frappe.db.sql("UPDATE `tabFreight Booking` SET container=NULL WHERE container=%s", cname)
			frappe.db.sql("DELETE FROM `tabImport Container Item` WHERE parent=%s", cname)
			frappe.db.sql("DELETE FROM `tabImport Container` WHERE name=%s", cname)

		frappe.db.commit()

	# 2. Count remaining containers in DB
	remaining_count = frappe.db.count("Import Container") if not dry_run else (len(all_db_containers) - deleted_count)

	# 3. Ensure Freight Bookings are linked via Commercial Invoice or Container
	if not dry_run:
		# Update container commercial invoice linkages where missing
		frappe.db.sql("""
			UPDATE `tabImport Container` c
			JOIN `tabFreight Booking` fb ON (fb.container = c.name OR fb.container = c.container_number)
			SET c.commercial_invoice = fb.commercial_invoice
			WHERE (c.commercial_invoice IS NULL OR c.commercial_invoice = '') AND fb.commercial_invoice IS NOT NULL
		""")

		# Update freight booking container linkages where missing
		frappe.db.sql("""
			UPDATE `tabFreight Booking` fb
			JOIN `tabImport Container` c ON c.commercial_invoice = fb.commercial_invoice
			SET fb.container = c.name
			WHERE fb.container IS NULL AND c.commercial_invoice IS NOT NULL
		""")

		frappe.db.commit()

	mode = "DRY-RUN" if dry_run else "APPLIED"

	print(f"\n========================================================")
	print(f"  CLEANUP & BIND 978 MASTER CONTAINERS ({mode})")
	print(f"========================================================")
	print(f"Total Master List Containers: {len(master_set)}")
	print(f"Non-Master Containers Deleted: {deleted_count}")
	print(f"Final Active Containers Count: {remaining_count}\n")

	return {
		"master_count": len(master_set),
		"deleted_count": deleted_count,
		"final_active_count": remaining_count,
	}
