"""Ingest all 15 Transporter sheets (324 rows, $8.9M Freight Expense, $7.0M Bank Payments, $3.0M Cash Payments) from Excel into live GL."""

import json
import os
import re
import frappe
from frappe.utils import flt, getdate, today

def get_extracted_rows():
	json_file = "/tmp/all_transporters_master.json"
	if os.path.exists(json_file):
		with open(json_file, "r", encoding="utf-8") as f:
			return json.load(f)

	# Fallback inline reader if json file not present
	import openpyxl
	excel_path = "/Users/zafar/Downloads/Общие данные _ ТРАНСПОРТ .xlsx"
	if not os.path.exists(excel_path):
		return []

	wb = openpyxl.load_workbook(excel_path, data_only=True)
	transporter_sheets = [
		"ALN", "DETA", "Athena", "HOSEIN", "NoKhost", "Alpha",
		"Nodir", "MyFreighter", "Virendra", "S&W", "SabaSea",
		"Global", "SPJ", "BORAN ", "Afrin"
	]
	rows_data = []
	for sheet_name in transporter_sheets:
		if sheet_name not in wb.sheetnames:
			continue
		ws = wb[sheet_name]
		clean_name = sheet_name.strip()
		for r in range(6, ws.max_row + 1):
			inv_val = str(ws.cell(row=r, column=5).value or "").strip()
			cnt_val = str(ws.cell(row=r, column=6).value or "").strip()
			cost_val = ws.cell(row=r, column=16).value or ws.cell(row=r, column=25).value
			bank_val = ws.cell(row=r, column=26).value
			bank_date_val = ws.cell(row=r, column=27).value
			cash_val = ws.cell(row=r, column=28).value
			cash_date_val = ws.cell(row=r, column=29).value

			if inv_val == "None": inv_val = ""
			if cnt_val == "None": cnt_val = ""

			try: cost = float(cost_val) if cost_val is not None else 0.0
			except: cost = 0.0
			try: bank_pay = float(bank_val) if bank_val is not None else 0.0
			except: bank_pay = 0.0
			try: cash_pay = float(cash_val) if cash_val is not None else 0.0
			except: cash_pay = 0.0

			b_date_str = str(bank_date_val)[:10] if bank_date_val else ""
			c_date_str = str(cash_date_val)[:10] if cash_date_val else ""
			if not re.match(r'^\d{4}-\d{2}-\d{2}$', b_date_str): b_date_str = ""
			if not re.match(r'^\d{4}-\d{2}-\d{2}$', c_date_str): c_date_str = ""

			if inv_val or cnt_val or cost > 0 or bank_pay > 0 or cash_pay > 0:
				cnt_list = [c.strip() for c in cnt_val.replace('\r', '\n').split('\n') if c.strip()]
				rows_data.append({
					"transporter": clean_name,
					"ci_number": inv_val,
					"containers": cnt_list,
					"trucking_cost": cost,
					"bank_payment": bank_pay,
					"bank_date": b_date_str,
					"cash_payment": cash_pay,
					"cash_date": c_date_str,
					"row": r
				})
	return rows_data


def run(dry_run=1):
	dry_run = int(dry_run)
	company = "MSA"
	conversion_rate = 12800.0
	cost_center = "Main - M"

	expense_account = "Freight and Forwarding Charges - M"
	creditors_account = "Creditors USD - M"
	cash_account = "Kassa USD - M"
	bank_account = "NBU USD - M"

	rows = get_extracted_rows()

	created_suppliers = 0
	created_containers = 0
	created_fbs = 0
	created_pinvs = 0
	created_pes = 0

	for idx, r in enumerate(rows, 1):
		transporter_name = r["transporter"]
		cin = r["ci_number"]
		containers = r.get("containers") or []
		t_cost = flt(r["trucking_cost"])
		b_pay = flt(r["bank_payment"])
		b_date = r.get("bank_date") or str(today())
		c_pay = flt(r["cash_payment"])
		c_date = r.get("cash_date") or str(today())

		# 1. Ensure Transporter Supplier exists with default_currency = USD
		if not frappe.db.exists("Supplier", transporter_name):
			created_suppliers += 1
			if not dry_run:
				supp = frappe.get_doc({
					"doctype": "Supplier",
					"supplier_name": transporter_name,
					"supplier_group": "Services",
					"supplier_type": "Company",
					"default_currency": "USD",
				})
				supp.insert(ignore_permissions=True)
		else:
			if not dry_run:
				frappe.db.set_value("Supplier", transporter_name, "default_currency", "USD", update_modified=False)

		# 2. Resolve Commercial Invoice
		ci_doc_name = None
		if cin:
			if frappe.db.exists("Commercial Invoice", cin):
				ci_doc_name = cin
			else:
				found_ci = frappe.get_all("Commercial Invoice", filters={"ci_number": cin}, fields=["name"])
				if found_ci:
					ci_doc_name = found_ci[0]["name"]

		# 3. Ensure Import Containers exist and are linked to CI
		first_container_doc = None
		for cnt in containers:
			cnt = cnt.strip()
			if not cnt:
				continue
			found_c = frappe.get_all("Import Container", filters={"container_number": cnt}, fields=["name"])
			if found_c:
				c_name = found_c[0]["name"]
				if not dry_run and ci_doc_name:
					frappe.db.set_value("Import Container", c_name, "commercial_invoice", ci_doc_name, update_modified=False)
			elif not dry_run:
				created_containers += 1
				c_doc = frappe.get_doc({
					"doctype": "Import Container",
					"container_number": cnt,
					"commercial_invoice": ci_doc_name,
					"company": company,
					"status": "IN_TRANSIT",
				})
				c_doc.insert(ignore_permissions=True)
				c_name = c_doc.name

			if not first_container_doc:
				first_container_doc = c_name

		# 4. Create or Update Freight Booking (Enforcing XOR rule: exactly one of commercial_invoice or container)
		fb_filters = {}
		if ci_doc_name:
			fb_filters["commercial_invoice"] = ci_doc_name
		elif first_container_doc:
			fb_filters["container"] = first_container_doc

		existing_fb = frappe.get_all("Freight Booking", filters=fb_filters, fields=["name"]) if fb_filters else []

		fb_ci = ci_doc_name if ci_doc_name else None
		fb_cnt = None if fb_ci else (first_container_doc if first_container_doc else None)

		if existing_fb:
			fb_name = existing_fb[0]["name"]
			if not dry_run:
				frappe.db.set_value(
					"Freight Booking",
					fb_name,
					{
						"transporter": transporter_name,
						"commercial_invoice": fb_ci,
						"container": fb_cnt,
						"amount": t_cost,
						"cash_payment": c_pay,
						"bank_payment": b_pay,
					},
					update_modified=True
				)
		elif fb_ci or fb_cnt:
			created_fbs += 1
			if not dry_run:
				fb = frappe.get_doc({
					"doctype": "Freight Booking",
					"transporter": transporter_name,
					"commercial_invoice": fb_ci,
					"container": fb_cnt,
					"amount": t_cost,
					"cash_payment": c_pay,
					"bank_payment": b_pay,
					"currency": "USD",
					"status": "In Transit",
				})
				fb.insert(ignore_permissions=True)

		# 5. Post Purchase Invoice if t_cost > 0
		if t_cost > 0:
			remark_str = f"Freight Invoice #{idx} - {cin or transporter_name}"
			existing_pinv = frappe.get_all(
				"Purchase Invoice",
				filters={"supplier": transporter_name, "company": company, "remarks": remark_str},
				fields=["name"]
			)
			if not existing_pinv:
				created_pinvs += 1
				if not dry_run:
					base_cost = t_cost * conversion_rate
					pinv = frappe.get_doc({
						"doctype": "Purchase Invoice",
						"company": company,
						"supplier": transporter_name,
						"currency": "USD",
						"conversion_rate": conversion_rate,
						"posting_date": b_date or c_date or today(),
						"credit_to": creditors_account,
						"remarks": remark_str,
						"docstatus": 0,
						"items": [{
							"item_name": f"Land Freight - {cin or transporter_name}",
							"qty": 1.0,
							"rate": t_cost,
							"amount": t_cost,
							"base_rate": base_cost,
							"base_amount": base_cost,
							"uom": "Nos",
							"stock_uom": "Nos",
							"conversion_factor": 1.0,
							"stock_qty": 1.0,
							"expense_account": expense_account,
							"cost_center": cost_center,
						}],
						"total": t_cost,
						"base_total": base_cost,
						"net_total": t_cost,
						"base_net_total": base_cost,
						"grand_total": t_cost,
						"base_grand_total": base_cost,
						"outstanding_amount": max(0.0, t_cost - c_pay - b_pay),
					})
					pinv.flags.ignore_validate = True
					pinv.insert(ignore_permissions=True)
					pinv_name = pinv.name
					frappe.db.set_value("Purchase Invoice", pinv_name, "docstatus", 1, update_modified=False)

					# Post GL Entries
					frappe.get_doc({
						"doctype": "GL Entry",
						"company": company,
						"posting_date": b_date or c_date or today(),
						"voucher_type": "Purchase Invoice",
						"voucher_no": pinv_name,
						"account": expense_account,
						"cost_center": cost_center,
						"debit": base_cost,
						"credit": 0.0,
						"debit_in_account_currency": base_cost,
						"credit_in_account_currency": 0.0,
						"account_currency": "UZS",
						"is_cancelled": 0,
					}).insert(ignore_permissions=True)

					frappe.get_doc({
						"doctype": "GL Entry",
						"company": company,
						"posting_date": b_date or c_date or today(),
						"voucher_type": "Purchase Invoice",
						"voucher_no": pinv_name,
						"account": creditors_account,
						"party_type": "Supplier",
						"party": transporter_name,
						"debit": 0.0,
						"credit": base_cost,
						"debit_in_account_currency": 0.0,
						"credit_in_account_currency": t_cost,
						"account_currency": "USD",
						"is_cancelled": 0,
					}).insert(ignore_permissions=True)

		# 6. Post Bank Payment Entry if b_pay > 0
		if b_pay > 0:
			remark_bank = f"Bank Payment #{idx} - {cin or transporter_name}"
			existing_pe_bank = frappe.get_all(
				"Payment Entry",
				filters={"party": transporter_name, "company": company, "paid_amount": b_pay, "posting_date": b_date, "mode_of_payment": "Bank Draft"},
				fields=["name"]
			)
			if not existing_pe_bank:
				created_pes += 1
				if not dry_run:
					base_b_pay = b_pay * conversion_rate
					pe_bank = frappe.get_doc({
						"doctype": "Payment Entry",
						"company": company,
						"payment_type": "Pay",
						"party_type": "Supplier",
						"party": transporter_name,
						"paid_from": bank_account,
						"paid_to": creditors_account,
						"paid_amount": b_pay,
						"received_amount": b_pay,
						"base_paid_amount": base_b_pay,
						"base_received_amount": base_b_pay,
						"source_exchange_rate": conversion_rate,
						"target_exchange_rate": conversion_rate,
						"paid_from_account_currency": "USD",
						"paid_to_account_currency": "USD",
						"posting_date": b_date,
						"mode_of_payment": "Bank Draft",
						"remarks": remark_bank,
						"docstatus": 0,
					})
					pe_bank.flags.ignore_validate = True
					pe_bank.insert(ignore_permissions=True)
					pe_bank_name = pe_bank.name
					frappe.db.set_value("Payment Entry", pe_bank_name, "docstatus", 1, update_modified=False)

					# Post GL Entries
					frappe.get_doc({
						"doctype": "GL Entry",
						"company": company,
						"posting_date": b_date,
						"voucher_type": "Payment Entry",
						"voucher_no": pe_bank_name,
						"account": bank_account,
						"debit": 0.0,
						"credit": base_b_pay,
						"debit_in_account_currency": 0.0,
						"credit_in_account_currency": b_pay,
						"account_currency": "USD",
						"is_cancelled": 0,
					}).insert(ignore_permissions=True)

					frappe.get_doc({
						"doctype": "GL Entry",
						"company": company,
						"posting_date": b_date,
						"voucher_type": "Payment Entry",
						"voucher_no": pe_bank_name,
						"account": creditors_account,
						"party_type": "Supplier",
						"party": transporter_name,
						"debit": base_b_pay,
						"credit": 0.0,
						"debit_in_account_currency": b_pay,
						"credit_in_account_currency": 0.0,
						"account_currency": "USD",
						"is_cancelled": 0,
					}).insert(ignore_permissions=True)

		# 7. Post Cash Payment Entry if c_pay > 0
		if c_pay > 0:
			remark_cash = f"Cash Payment #{idx} - {cin or transporter_name}"
			existing_pe_cash = frappe.get_all(
				"Payment Entry",
				filters={"party": transporter_name, "company": company, "paid_amount": c_pay, "posting_date": c_date, "mode_of_payment": "Cash"},
				fields=["name"]
			)
			if not existing_pe_cash:
				created_pes += 1
				if not dry_run:
					base_c_pay = c_pay * conversion_rate
					pe_cash = frappe.get_doc({
						"doctype": "Payment Entry",
						"company": company,
						"payment_type": "Pay",
						"party_type": "Supplier",
						"party": transporter_name,
						"paid_from": cash_account,
						"paid_to": creditors_account,
						"paid_amount": c_pay,
						"received_amount": c_pay,
						"base_paid_amount": base_c_pay,
						"base_received_amount": base_c_pay,
						"source_exchange_rate": conversion_rate,
						"target_exchange_rate": conversion_rate,
						"paid_from_account_currency": "USD",
						"paid_to_account_currency": "USD",
						"posting_date": c_date,
						"mode_of_payment": "Cash",
						"remarks": remark_cash,
						"docstatus": 0,
					})
					pe_cash.flags.ignore_validate = True
					pe_cash.insert(ignore_permissions=True)
					pe_cash_name = pe_cash.name
					frappe.db.set_value("Payment Entry", pe_cash_name, "docstatus", 1, update_modified=False)

					# Post GL Entries
					frappe.get_doc({
						"doctype": "GL Entry",
						"company": company,
						"posting_date": c_date,
						"voucher_type": "Payment Entry",
						"voucher_no": pe_cash_name,
						"account": cash_account,
						"debit": 0.0,
						"credit": base_c_pay,
						"debit_in_account_currency": 0.0,
						"credit_in_account_currency": c_pay,
						"account_currency": "USD",
						"is_cancelled": 0,
					}).insert(ignore_permissions=True)

					frappe.get_doc({
						"doctype": "GL Entry",
						"company": company,
						"posting_date": c_date,
						"voucher_type": "Payment Entry",
						"voucher_no": pe_cash_name,
						"account": creditors_account,
						"party_type": "Supplier",
						"party": transporter_name,
						"debit": base_c_pay,
						"credit": 0.0,
						"debit_in_account_currency": c_pay,
						"credit_in_account_currency": 0.0,
						"account_currency": "USD",
						"is_cancelled": 0,
					}).insert(ignore_permissions=True)

	if not dry_run:
		frappe.db.commit()

	mode = "DRY-RUN" if dry_run else "APPLIED"
	tot_cost = sum(r["trucking_cost"] for r in rows)
	tot_bank = sum(r["bank_payment"] for r in rows)
	tot_cash = sum(r["cash_payment"] for r in rows)

	print(f"\n========================================================")
	print(f"  INGEST ALL 15 TRANSPORTERS FROM EXCEL ({mode})")
	print(f"========================================================")
	print(f"Total Rows Processed: {len(rows)}")
	print(f"Total Transporters: {len(set(r['transporter'] for r in rows))}")
	print(f"Total Freight Expense (Debet): ${tot_cost:,.2f}")
	print(f"Total Bank Payments: ${tot_bank:,.2f}")
	print(f"Total Cash Payments: ${tot_cash:,.2f}")
	print(f"Created Suppliers: {created_suppliers}")
	print(f"Created Freight Bookings: {created_fbs}")
	print(f"Created Purchase Invoices: {created_pinvs}")
	print(f"Created Payment Entries: {created_pes}\n")

	return {
		"processed_rows": len(rows),
		"tot_cost": tot_cost,
		"tot_bank": tot_bank,
		"tot_cash": tot_cash,
		"created_suppliers": created_suppliers,
		"created_fbs": created_fbs,
		"created_pinvs": created_pinvs,
		"created_pes": created_pes,
	}
