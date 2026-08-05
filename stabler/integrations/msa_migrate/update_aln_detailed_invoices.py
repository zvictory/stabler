"""Update ALN Purchase Invoices with detailed Itemized Breakdown: Base Freight (P=$2,774,600.00), Demurrage Charges (R=$9,221.80), and Additional Trucking Costs (W=$16,100.00), matching Z column Total Cost ($2,961,365.80)."""

import json
import os
import frappe
from frappe.utils import flt, today

def get_aln_breakdown():
	json_file = "/tmp/aln_cost_breakdown.json"
	if os.path.exists(json_file):
		with open(json_file, "r", encoding="utf-8") as f:
			return json.load(f)

	# Inline extraction fallback
	import openpyxl
	excel_path = "/Users/zafar/Downloads/Общие данные _ ТРАНСПОРТ .xlsx"
	wb = openpyxl.load_workbook(excel_path, data_only=True)
	ws = wb["ALN"]
	rows_data = []
	for r in range(6, ws.max_row + 1):
		inv = str(ws.cell(row=r, column=5).value or "").strip()
		cnt = str(ws.cell(row=r, column=6).value or "").strip()
		if inv == "None": inv = ""
		if cnt == "None": cnt = ""

		base_p = ws.cell(row=r, column=16).value or 0.0
		dem_r = ws.cell(row=r, column=18).value or 0.0
		stor_s = ws.cell(row=r, column=19).value or 0.0
		add_w = ws.cell(row=r, column=23).value or 0.0
		tot_z = ws.cell(row=r, column=26).value or 0.0

		try: base_p = float(base_p)
		except: base_p = 0.0
		try: dem_r = float(dem_r)
		except: dem_r = 0.0
		try: stor_s = float(stor_s)
		except: stor_s = 0.0
		try: add_w = float(add_w)
		except: add_w = 0.0
		try: tot_z = float(tot_z)
		except: tot_z = 0.0

		if inv or cnt or tot_z > 0 or base_p > 0:
			cnt_list = [c.strip() for c in cnt.replace('\r', '\n').split('\n') if c.strip()]
			rows_data.append({
				"row": r,
				"invoice": inv,
				"containers": cnt_list,
				"base_trucking": base_p,
				"demurrage": dem_r,
				"storage": stor_s,
				"additional_trucking": add_w,
				"total_cost": tot_z if tot_z > 0 else (base_p + dem_r + stor_s + add_w)
			})
	return rows_data


def run(dry_run=1):
	dry_run = int(dry_run)
	company = "MSA"
	transporter = "ALN"
	conversion_rate = 12800.0
	cost_center = "Main - M"

	expense_account = "Freight and Forwarding Charges - M"
	creditors_account = "Creditors USD - M"

	rows = get_aln_breakdown()

	if not dry_run:
		# Delete old Purchase Invoices for ALN
		existing_pinvs = frappe.get_all("Purchase Invoice", filters={"supplier": transporter, "company": company}, fields=["name"])
		for pinv in existing_pinvs:
			frappe.db.sql("DELETE FROM `tabGL Entry` WHERE voucher_type='Purchase Invoice' AND voucher_no=%s", pinv["name"])
			frappe.db.sql("DELETE FROM `tabPurchase Invoice Item` WHERE parent=%s", pinv["name"])
			frappe.db.sql("DELETE FROM `tabPurchase Invoice` WHERE name=%s", pinv["name"])
		frappe.db.commit()

	created_pinvs = 0
	tot_base = 0.0
	tot_dem = 0.0
	tot_add = 0.0
	tot_grand = 0.0

	for idx, r in enumerate(rows, 1):
		cin = r["invoice"]
		containers = r.get("containers") or []
		base_p = flt(r["base_trucking"])
		dem_r = flt(r["demurrage"])
		add_w = flt(r["additional_trucking"])
		tot_z = flt(r["total_cost"])

		if tot_z <= 0:
			tot_z = base_p + dem_r + add_w

		tot_base += base_p
		tot_dem += dem_r
		tot_add += add_w
		tot_grand += tot_z

		if tot_z > 0:
			created_pinvs += 1
			if not dry_run:
				base_tot_z = tot_z * conversion_rate
				remark_str = f"ALN Detailed Freight Invoice #{idx} - {cin or 'Batch'}"

				items = []
				if base_p > 0:
					b_rate_uzs = base_p * conversion_rate
					items.append({
						"item_name": f"Land Freight - {cin or 'Batch'}",
						"qty": 1.0,
						"rate": base_p,
						"amount": base_p,
						"base_rate": b_rate_uzs,
						"base_amount": b_rate_uzs,
						"uom": "Nos",
						"stock_uom": "Nos",
						"conversion_factor": 1.0,
						"stock_qty": 1.0,
						"expense_account": expense_account,
						"cost_center": cost_center,
					})
				if dem_r > 0:
					d_rate_uzs = dem_r * conversion_rate
					items.append({
						"item_name": f"Demurrage Charge at BND Port - {cin or 'Batch'}",
						"qty": 1.0,
						"rate": dem_r,
						"amount": dem_r,
						"base_rate": d_rate_uzs,
						"base_amount": d_rate_uzs,
						"uom": "Nos",
						"stock_uom": "Nos",
						"conversion_factor": 1.0,
						"stock_qty": 1.0,
						"expense_account": expense_account,
						"cost_center": cost_center,
					})
				if add_w > 0:
					a_rate_uzs = add_w * conversion_rate
					items.append({
						"item_name": f"Additional Trucking Cost - {cin or 'Batch'}",
						"qty": 1.0,
						"rate": add_w,
						"amount": add_w,
						"base_rate": a_rate_uzs,
						"base_amount": a_rate_uzs,
						"uom": "Nos",
						"stock_uom": "Nos",
						"conversion_factor": 1.0,
						"stock_qty": 1.0,
						"expense_account": expense_account,
						"cost_center": cost_center,
					})

				pinv = frappe.get_doc({
					"doctype": "Purchase Invoice",
					"company": company,
					"supplier": transporter,
					"currency": "USD",
					"conversion_rate": conversion_rate,
					"posting_date": today(),
					"credit_to": creditors_account,
					"remarks": remark_str,
					"docstatus": 0,
					"items": items,
					"total": tot_z,
					"base_total": base_tot_z,
					"net_total": tot_z,
					"base_net_total": base_tot_z,
					"grand_total": tot_z,
					"base_grand_total": base_tot_z,
					"outstanding_amount": tot_z,
				})
				pinv.flags.ignore_validate = True
				pinv.insert(ignore_permissions=True)
				pinv_name = pinv.name
				frappe.db.set_value("Purchase Invoice", pinv_name, "docstatus", 1, update_modified=False)

				# Post GL Entries
				frappe.get_doc({
					"doctype": "GL Entry",
					"company": company,
					"posting_date": today(),
					"voucher_type": "Purchase Invoice",
					"voucher_no": pinv_name,
					"account": expense_account,
					"cost_center": cost_center,
					"debit": base_tot_z,
					"credit": 0.0,
					"debit_in_account_currency": base_tot_z,
					"credit_in_account_currency": 0.0,
					"account_currency": "UZS",
					"is_cancelled": 0,
				}).insert(ignore_permissions=True)

				frappe.get_doc({
					"doctype": "GL Entry",
					"company": company,
					"posting_date": today(),
					"voucher_type": "Purchase Invoice",
					"voucher_no": pinv_name,
					"account": creditors_account,
					"party_type": "Supplier",
					"party": transporter,
					"debit": 0.0,
					"credit": base_tot_z,
					"debit_in_account_currency": 0.0,
					"credit_in_account_currency": tot_z,
					"account_currency": "USD",
					"is_cancelled": 0,
				}).insert(ignore_permissions=True)

	if not dry_run:
		frappe.db.commit()

	mode = "DRY-RUN" if dry_run else "APPLIED"

	print(f"\n========================================================")
	print(f"  UPDATE ALN DETAILED INVOICES BREAKDOWN ({mode})")
	print(f"========================================================")
	print(f"Total Rows Processed: {len(rows)}")
	print(f"Created Purchase Invoices: {created_pinvs}")
	print(f"Base Trucking Total (P): ${tot_base:,.2f}")
	print(f"Demurrage Total (R):     ${tot_dem:,.2f}")
	print(f"Additional Cost (W):     ${tot_add:,.2f}")
	print(f"GRAND TOTAL COST (Z):    ${tot_grand:,.2f}\n")

	return {
		"processed_rows": len(rows),
		"created_pinvs": created_pinvs,
		"tot_base": tot_base,
		"tot_dem": tot_dem,
		"tot_add": tot_add,
		"tot_grand": tot_grand,
	}
